"""Unit tests: schema mapping + canonical normalize (no browser)."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from material_price_audit.export_quotes import write_quote_result_workbook
from material_price_audit.models import Quote, QuoteSet
from material_price_audit.normalize import load_canonical_items
from material_price_audit.schema_map import map_sheet_by_rules
from material_price_audit.settings_store import UserSettings, load_settings, save_settings


def _make_weird_sheet(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "安装"
    # non-standard headers
    ws["A1"] = "某某项目询价表"
    ws["A3"] = "序号"
    ws["B3"] = "材料设备名称"
    ws["C3"] = "规格及型号"
    ws["D3"] = "计量单位"
    ws["E3"] = "工程量"
    ws["F3"] = "承包人报价(不含税)"
    ws["G3"] = "产地品牌"
    ws["B4"] = "热镀锌钢管"
    ws["C4"] = "DN100"
    ws["D4"] = "m"
    ws["E4"] = 120
    ws["F4"] = 85.5
    ws["G4"] = "国产"
    ws["B5"] = "海康威视 DS-KH6320-C1 室内机"
    ws["C5"] = ""
    ws["D5"] = "台"
    ws["E5"] = 2
    ws["F5"] = 1200
    wb.save(path)


def test_rule_map_nonstandard_headers(tmp_path: Path):
    xlsx = tmp_path / "weird.xlsx"
    _make_weird_sheet(xlsx)
    wb = openpyxl.load_workbook(xlsx)
    schema = map_sheet_by_rules(wb["安装"], "安装")
    assert schema is not None
    assert schema.header_row == 3
    roles = schema.roles()
    assert roles.get("name") == 2
    assert roles.get("spec") == 3
    assert roles.get("unit") == 4
    assert roles.get("qty") == 5
    assert roles.get("submit_price") == 6


def test_load_canonical_and_export(tmp_path: Path):
    from material_price_audit.models import SheetSchema, WorkbookSchema, ColumnMap

    xlsx = tmp_path / "weird.xlsx"
    _make_weird_sheet(xlsx)
    wb = openpyxl.load_workbook(xlsx)
    sch = map_sheet_by_rules(wb["安装"], "安装")
    assert sch
    wschema = WorkbookSchema(file_fingerprint="t", sheets=[sch])
    items = load_canonical_items(xlsx, wschema)
    assert len(items) >= 2
    assert items[0].name
    assert items[0].search_queries
    # glued model in name
    hik = next((i for i in items if "6320" in i.text or "海康" in i.name), None)
    assert hik is not None
    assert hik.search_queries

    qmap = {
        items[0].id: QuoteSet(
            item_id=items[0].id,
            status="full_k",
            quotes=[
                Quote(rank=1, price=80, platform="guangcai", title="镀锌管 DN100", url="https://example.com/1", match_level="strict", match_score=0.9, price_ex_tax=70.8),
                Quote(rank=2, price=82, platform="huixun", title="热镀锌 DN100", url="https://example.com/2", match_level="approximate", match_score=0.7, price_ex_tax=72.57),
                Quote(rank=3, price=79, platform="lingcai", title="钢管 DN100", url="https://example.com/3", match_level="approximate", match_score=0.65, price_ex_tax=69.91),
            ],
        )
    }
    out = tmp_path / "result.xlsx"
    stats = write_quote_result_workbook(xlsx, out, items, qmap, k=3)
    assert out.exists()
    assert stats["full_k"] == 1
    rwb = openpyxl.load_workbook(out)
    assert "询价比价结果" in rwb.sheetnames


def test_settings_roundtrip(tmp_path: Path):
    s = UserSettings(platforms_enabled=["guangcai", "huixun"], quotes_per_item=3)
    save_settings(tmp_path, s)
    s2 = load_settings(tmp_path)
    assert s2.platforms_enabled == ["guangcai", "huixun"]
    assert s2.quotes_per_item == 3
