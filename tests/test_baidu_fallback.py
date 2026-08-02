"""百度全网兜底：触发条件、预算、分类、去重、不采摘要价、不编价。"""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import pytest

from material_price_audit.adapters import baidu_fallback as bf
from material_price_audit.models import CanonicalItem, Quote, QuoteSet
from material_price_audit.settings_store import UserSettings


def _item(name="球墨铸铁管", spec="DN150 PN10", **kw) -> CanonicalItem:
    return CanonicalItem(
        id=kw.get("id", "t1"),
        sheet="s1",
        row=2,
        name=name,
        spec=spec,
        brand=kw.get("brand", ""),
        unit=kw.get("unit", "m"),
    )


# —— 1. 原平台已满 K 不查百度 ——
def test_full_k_does_not_trigger_baidu():
    assert (
        bf.should_trigger_baidu(
            formal_quote_count=3,
            k=3,
            baidu_already_done=False,
            baidu_enabled=True,
        )
        is False
    )


def test_run_baidu_skips_when_full_k():
    item = _item()
    with patch.object(bf, "fetch_baidu_serp") as mock_fetch:
        res = bf.run_baidu_fallback(
            item,
            page=None,
            formal_quote_count=3,
            k=3,
            baidu_enabled=True,
            already_done=False,
        )
        mock_fetch.assert_not_called()
    assert res.skipped_reason == "full_k"
    assert res.web_refs == []
    assert res.supplier_leads == []


# —— 2. 不足 K 时只触发一次 ——
def test_partial_triggers_once_already_done_blocks():
    assert (
        bf.should_trigger_baidu(
            formal_quote_count=1,
            k=3,
            baidu_already_done=False,
            baidu_enabled=True,
        )
        is True
    )
    assert (
        bf.should_trigger_baidu(
            formal_quote_count=1,
            k=3,
            baidu_already_done=True,
            baidu_enabled=True,
        )
        is False
    )


def test_run_baidu_skips_when_already_done():
    with patch.object(bf, "fetch_baidu_serp") as mock_fetch:
        res = bf.run_baidu_fallback(
            _item(),
            formal_quote_count=0,
            k=3,
            already_done=True,
            baidu_enabled=True,
        )
        mock_fetch.assert_not_called()
    assert res.skipped_reason == "already_done"


# —— 3. 每条材料最多 2 个查询词 ——
def test_max_two_queries():
    qs = bf.build_baidu_queries(
        "球墨铸铁管（承插）",
        "DN150 PN10 壁厚 6mm",
        max_n=2,
    )
    assert len(qs) <= 2
    assert qs
    # 硬规格保留
    blob = " ".join(qs)
    assert "DN150" in blob or "150" in blob


def test_hard_spec_tokens_preserved():
    toks = bf.extract_hard_spec_tokens("钢管", "DN200 PN16 3kW 220V")
    joined = " ".join(toks).upper()
    assert "DN200" in joined or "200" in joined
    assert "PN16" in joined or "16" in joined


# —— 4. 搜索摘要价格不能采用 ——
def test_snippet_price_not_used_only_page_body():
    # extract_visible_prices 只看传入 text；run 里故意不传 snippet
    snippet = "市场价：¥99999 元 超低价"
    # 空 body → 无价
    prices = bf.extract_visible_prices_from_page("")
    assert prices == []
    # 仅摘要不应被 classify 当作来源价（我们用 body 抽价）
    body_no_price = "球墨铸铁管 DN150 产品介绍 厂家直销 联系电话13800138000"
    prices2 = bf.extract_visible_prices_from_page(body_no_price)
    assert prices2 == []
    # 摘要里的价若误传入 body 才有——验证提取需「价格」上下文
    prices3 = bf.extract_visible_prices_from_page(snippet)
    assert prices3  # 有上下文时可抽；但 run 流程不把 snippet 当 body


def test_classify_without_page_price_is_supplier_not_web_ref():
    item = _item()
    title = "球墨铸铁管 DN150 PN10 厂家"
    body = "球墨铸铁管 DN150 PN10 本公司生产销售，联系电话 13900001111 某某管业有限公司"
    kind, q, _ = bf.classify_source_match(
        item,
        title,
        body,
        prices=[],  # 无来源页数字价
        contact={"supplier": "某某管业有限公司", "phone": "13900001111", "contact": ""},
        url="https://example-pipe.com/p/1",
        quality="medium",
    )
    assert kind == "supplier_lead"
    assert q is not None
    assert q.price_role == "supplier_lead"
    assert float(q.price or 0) == 0.0


# —— 5. 原页名称规格完全匹配 + 真价 → 仅 web_reference ——
def test_full_match_with_price_is_web_reference_not_formal():
    item = _item()
    title = "球墨铸铁管 DN150 PN10"
    body = (
        "产品名称：球墨铸铁管 DN150 PN10。"
        "价格：128.5 元/米。厂家某某管业有限公司。"
    )
    prices = bf.extract_visible_prices_from_page(body)
    assert prices and prices[0][0] == pytest.approx(128.5)
    kind, q, _ = bf.classify_source_match(
        item,
        title,
        body,
        prices=prices,
        contact={"supplier": "某某管业", "phone": "", "contact": ""},
        url="https://maker.example.com/product/1",
        quality="high",
    )
    assert kind == "web_reference"
    assert q is not None
    assert q.price_role == "web_reference"
    assert q.price == pytest.approx(128.5)
    assert q.platform == "baidu_web"


# —— 6. DN 冲突必须拒绝 ——
def test_dn_conflict_rejected():
    item = _item(name="球墨铸铁管", spec="DN150 PN10")
    title = "球墨铸铁管 DN200 PN10"
    body = "球墨铸铁管 DN200 PN10 价格：99 元"
    prices = [(99.0, "价格：99 元")]
    kind, q, _ = bf.classify_source_match(
        item,
        title,
        body,
        prices=prices,
        contact={},
        url="https://shop.example.com/x",
        quality="medium",
    )
    assert kind == "reject"
    assert q is None


# —— 7. 无价格有供应商 → supplier_lead ——
def test_no_price_with_supplier_is_lead():
    item = _item()
    title = "球墨铸铁管 DN150 生产厂家"
    body = (
        "球墨铸铁管 DN150 PN10 规格齐全。"
        "公司名称：华通管材有限公司 电话：0571-88886666"
    )
    contact = bf.extract_page_contact(body)
    kind, q, _ = bf.classify_source_match(
        item,
        title,
        body,
        prices=[],
        contact=contact,
        url="https://huatong-pipe.com/about",
        quality="medium",
    )
    assert kind == "supplier_lead"
    assert q is not None
    assert q.price == 0.0
    assert q.price_role == "supplier_lead"


# —— 8. 重复 URL / 查询词不重复访问 ——
def test_filter_dedupes_url_and_title():
    hits = [
        bf.SerpHit("球墨铸铁管 DN150", "https://a.com/p1", "", 1),
        bf.SerpHit("球墨铸铁管 DN150", "https://a.com/p1?from=baidu", "", 2),
        bf.SerpHit("球墨铸铁管 DN150 副本", "https://b.com/p2", "", 3),
        bf.SerpHit("垃圾知道", "https://zhidao.baidu.com/q/1", "", 4),
    ]
    ranked = bf.filter_and_rank_hits(hits, max_n=5)
    urls = [h.url for h in ranked]
    assert "zhidao.baidu.com" not in " ".join(urls)
    # p1 去重
    assert sum(1 for u in urls if "a.com/p1" in u) == 1


def test_run_baidu_does_not_reopen_same_url(tmp_path: Path):
    item = _item()
    serp = [
        bf.SerpHit("球墨铸铁管 DN150 PN10", "https://dup.example.com/x", "介绍", 1),
        bf.SerpHit("球墨铸铁管 DN150", "https://dup.example.com/x#frag", "介绍2", 2),
    ]
    page_html = (
        "<html><head><title>球墨铸铁管 DN150 PN10</title></head><body>"
        + ("产品详情页 " * 20)
        + "球墨铸铁管 DN150 PN10 价格：55 元 "
        "华通管材有限公司 电话13800138000"
        + (" 规格说明 " * 20)
        + "</body></html>"
    )
    open_count = {"n": 0}

    def fake_fetch(page, query, **kw):
        return serp, False, "live"

    def fake_http(page, url, timeout_ms=20000):
        open_count["n"] += 1
        return url, page_html

    with (
        patch.object(bf, "fetch_baidu_serp", side_effect=fake_fetch),
        patch.object(bf, "_http_get", side_effect=fake_http),
        patch.object(bf, "build_baidu_queries", return_value=["球墨铸铁管 DN150"]),
    ):
        res = bf.run_baidu_fallback(
            item,
            formal_quote_count=0,
            k=3,
            max_queries=2,
            max_pages=5,
            root=tmp_path,
        )
    # 相同规范化 URL 只打开一次
    assert open_count["n"] == 1
    assert len(res.web_refs) + len(res.supplier_leads) >= 1


# —— 9. 百度失败后可正常生成 RFQ ——
def test_baidu_failure_still_allows_rfq(tmp_path: Path):
    from material_price_audit.export_quotes import export_rfq_from_quotes

    item = _item()
    # baidu 失败：run 吞异常并返回空
    with patch.object(bf, "fetch_baidu_serp", side_effect=RuntimeError("network")):
        res = bf.run_baidu_fallback(
            item, formal_quote_count=0, k=3, root=tmp_path
        )
    assert res.web_refs == []
    qset = QuoteSet(
        item_id=item.id,
        quotes=[],
        status="no_match",
        error="没查到",
        supplier_leads=res.supplier_leads,
        web_refs=res.web_refs,
    )
    out = tmp_path / "rfq.xlsx"
    n = export_rfq_from_quotes([item], {item.id: qset}, out, k=3)
    assert n == 1
    assert out.exists()


def test_rfq_includes_supplier_lead_columns(tmp_path: Path):
    from material_price_audit.export_quotes import export_rfq_from_quotes
    import openpyxl

    item = _item()
    lead = Quote(
        rank=1,
        price=0.0,
        platform="baidu_web",
        title="线索",
        url="https://sup.example.com",
        price_role="supplier_lead",
        supplier="华通管材有限公司",
        phone="13800138000",
    )
    qset = QuoteSet(
        item_id=item.id,
        status="need_review",
        supplier_leads=[lead],
    )
    out = tmp_path / "rfq2.xlsx"
    export_rfq_from_quotes([item], {item.id: qset}, out, k=3)
    wb = openpyxl.load_workbook(out)
    assert "供应商线索" in wb.sheetnames
    ws2 = wb["供应商线索"]
    assert ws2.cell(2, 5).value == "华通管材有限公司"


# —— 10. 全过程不增加无预算的 LLM 调用 ——
def test_baidu_path_has_no_llm_import_or_call(tmp_path: Path):
    item = _item()
    serp = [bf.SerpHit("球墨铸铁管 DN150", "https://ok.example.com/p", "", 1)]
    html = (
        "<html><body>"
        + ("详情 " * 30)
        + "球墨铸铁管 DN150 PN10 价格：88 元 某某有限公司"
        + (" 说明 " * 30)
        + "</body></html>"
    )

    with (
        patch.object(bf, "fetch_baidu_serp", return_value=(serp, False, "live")),
        patch.object(bf, "_http_get", return_value=("https://ok.example.com/p", html)),
        patch.object(bf, "build_baidu_queries", return_value=["球墨铸铁管 DN150"]),
        patch.dict("sys.modules", {"material_price_audit.llm_agent": None}),
    ):
        # 即使 llm 模块不可用也应完成
        res = bf.run_baidu_fallback(
            item, formal_quote_count=0, k=2, root=tmp_path
        )
    # 不依赖 LLM
    assert res.skipped_reason == ""
    assert isinstance(res.web_refs, list)


def test_source_quality_blocks_baidu_know():
    assert bf.source_quality_for_url("https://zhidao.baidu.com/q/1") == "block"
    assert bf.source_quality_for_url("https://wenku.baidu.com/x") == "block"
    assert bf.source_quality_for_url("https://passport.baidu.com/v2/?login") == "block"
    assert bf.source_quality_for_url("https://www.baidu.com/baidu.php?url=x") == "block"
    assert bf.source_quality_for_url("https://ccgp.gov.cn/x") in ("high", "medium")


def test_empty_non_contact_page_is_not_supplier_lead():
    item = _item()
    kind, q, _ = bf.classify_source_match(
        item,
        "球墨铸铁管 DN150 PN10 商品页",
        "球墨铸铁管 DN150 PN10 产品展示，没有厂家和联系方式",
        prices=[],
        contact={"supplier": "", "phone": "", "contact": ""},
        url="https://shop.example.com/item/1",
        quality="medium",
    )
    assert kind in ("reject", "alias_only")
    assert q is None


def test_final_baidu_login_redirect_is_discarded(tmp_path: Path):
    hit = bf.SerpHit(
        "球墨铸铁管 DN150 PN10 厂家",
        "https://maker.example.com/product/1",
        "",
        1,
    )
    html = "<html><body>" + ("登录百度账号 " * 40) + "</body></html>"
    with (
        patch.object(bf, "fetch_baidu_serp", return_value=([hit], False, "live")),
        patch.object(
            bf,
            "_http_get",
            return_value=("https://passport.baidu.com/v2/?login", html),
        ),
        patch.object(bf, "build_baidu_queries", return_value=["球墨铸铁管 DN150"]),
    ):
        res = bf.run_baidu_fallback(
            _item(), formal_quote_count=0, k=3, root=tmp_path
        )
    assert res.web_refs == []
    assert res.supplier_leads == []
    assert any(a.get("status") == "reject_non_source_page" for a in res.attempts)


def test_quoteset_web_refs_roundtrip():
    qs = QuoteSet(
        item_id="x",
        web_refs=[
            Quote(
                rank=1,
                price=12.0,
                platform="baidu_web",
                title="t",
                url="https://a.com",
                price_role="web_reference",
                source_quality="high",
            )
        ],
        supplier_leads=[
            Quote(
                rank=1,
                price=0.0,
                platform="baidu_web",
                title="s",
                url="https://b.com",
                price_role="supplier_lead",
            )
        ],
    )
    d = qs.to_dict()
    qs2 = QuoteSet.from_dict(d)
    assert len(qs2.web_refs) == 1
    assert qs2.web_refs[0].price_role == "web_reference"
    assert qs2.web_refs[0].source_quality == "high"
    assert len(qs2.supplier_leads) == 1


def test_settings_baidu_default_off_until_user_confirms():
    s = UserSettings.from_dict({})
    assert s.baidu_fallback_enabled is False
    # 迁移旧版默认 true：没有新版用户确认标记也必须关闭。
    legacy = UserSettings.from_dict({"baidu_fallback_enabled": True})
    assert legacy.baidu_fallback_enabled is False
    explicit = UserSettings.from_dict(
        {
            "baidu_fallback_enabled": True,
            "baidu_fallback_confirmed": True,
        }
    )
    assert explicit.baidu_fallback_enabled is True
    s2 = UserSettings.from_dict({"baidu_fallback_enabled": False})
    assert s2.baidu_fallback_enabled is False


def test_captcha_stops_baidu(tmp_path: Path):
    with (
        patch.object(
            bf, "fetch_baidu_serp", return_value=([], True, "captcha")
        ),
        patch.object(bf, "build_baidu_queries", return_value=["q1", "q2"]),
        patch.object(bf, "_http_get") as mock_open,
    ):
        res = bf.run_baidu_fallback(
            _item(), formal_quote_count=0, k=3, root=tmp_path
        )
        mock_open.assert_not_called()
    assert res.captcha is True
    assert res.web_refs == []


def test_disabled_baidu():
    assert (
        bf.should_trigger_baidu(
            formal_quote_count=0,
            k=3,
            baidu_already_done=False,
            baidu_enabled=False,
        )
        is False
    )


def test_serp_price_in_snippet_ignored_in_run(tmp_path: Path):
    """SERP snippet 含价格，但页面无价 → 不得变成 web_reference。"""
    item = _item()
    serp = [
        bf.SerpHit(
            "球墨铸铁管 DN150 PN10",
            "https://page.example.com/p",
            "特价 ¥1 元包邮",  # 摘要价
            1,
        )
    ]
    # 页面只有名称规格+厂家，无数字价
    html = (
        "<html><body>"
        + ("介绍 " * 30)
        + "球墨铸铁管 DN150 PN10 华通管材有限公司 "
        "联系电话13912345678 产品介绍"
        + (" 内容 " * 30)
        + "</body></html>"
    )
    with (
        patch.object(bf, "fetch_baidu_serp", return_value=(serp, False, "live")),
        patch.object(bf, "_http_get", return_value=("https://page.example.com/p", html)),
        patch.object(bf, "build_baidu_queries", return_value=["球墨铸铁管 DN150"]),
    ):
        res = bf.run_baidu_fallback(
            item, formal_quote_count=0, k=3, root=tmp_path
        )
    assert res.web_refs == []
    # 可进 supplier_lead
    assert all(q.price_role == "supplier_lead" for q in res.supplier_leads)
    assert all(float(q.price or 0) == 0 for q in res.supplier_leads)


def test_max_pages_budget(tmp_path: Path):
    item = _item()
    serp = [
        bf.SerpHit(f"球墨铸铁管 DN150 #{i}", f"https://site{i}.example.com/p", "", i)
        for i in range(1, 12)
    ]
    opens: list[str] = []

    def fake_http(page, url, timeout_ms=20000):
        opens.append(url)
        body = (
            f"球墨铸铁管 DN150 PN10 价格：{10 + len(opens)} 元 "
            f"厂{len(opens)}有限公司"
        )
        return (
            url,
            "<html><body>" + ("x " * 40) + body + (" y" * 40) + "</body></html>",
        )

    with (
        patch.object(bf, "fetch_baidu_serp", return_value=(serp, False, "live")),
        patch.object(bf, "_http_get", side_effect=fake_http),
        patch.object(bf, "build_baidu_queries", return_value=["球墨铸铁管 DN150"]),
    ):
        bf.run_baidu_fallback(
            item,
            formal_quote_count=0,
            k=3,
            max_pages=5,
            root=tmp_path,
        )
    assert len(opens) <= 5
