from pathlib import Path

import openpyxl

from material_price_audit.export_quotes import write_quote_result_workbook
from material_price_audit.inquiry import build_review_candidates
from material_price_audit.models import CanonicalItem, Quote, QuoteSet


def test_exact_name_candidate_is_kept_for_review_not_as_strict_quote():
    item = CanonicalItem(
        id="x",
        sheet="询价",
        row=3,
        name="8端口分控器",
        spec="脱机分控器，AC220V/8端口分控制器,各端口标准512通道",
        submit=960.0,
    )
    attempts = [
        {
            "platform": "guangcai",
            "url": "https://www.gldjc.com/scj/so.html?keyword=8端口分控器",
            "quotation_url": "https://example.com/quote.jpg",
            "price_tax": 3747.08,  # 远超报送 — 仍保留，标 above_submit
            "match_ok": False,
            "match_score": 0.4,
            "match_detail": "名称命中；规格缺少：电压 AC220V, 512通道, 脱机",
            "title": "控制器",
            "spec_seen": "品种 : 分控器 | 产品描述 : 8端口分控器",
            "supplier": "安徽富晟兴智能科技有限公司",
            "name_hit": True,
            "tax_mode": "tax_excl",
        },
        {
            "platform": "guangcai",
            "url": "https://www.gldjc.com/scj/so.html?keyword=分控器",
            "price_tax": 900.0,
            "match_ok": False,
            "match_score": 0.6,
            "match_detail": "名称命中；规格缺少：电压 AC220V, 脱机",
            "title": "分控器",
            "spec_seen": "8路独立信号数据输出，可同时带载DMX512",
            "name_hit": True,
            "tax_mode": "tax_excl",
            "detail_url": "https://example.com/ok",
        },
    ]

    review = build_review_candidates(item, attempts, limit=3)
    assert len(review) >= 1
    # 分高的 900 应靠前；高价 3747 也保留
    prices = {r.price for r in review}
    assert 900.0 in prices
    assert review[0].match_level == "need_review"

    qset = QuoteSet(item_id=item.id, review_candidates=review, status="need_review")
    restored = QuoteSet.from_dict(qset.to_dict())
    assert restored.status == "need_review"
    assert restored.quotes == []
    assert restored.review_candidates[0].price > 0
    assert restored.review_candidates[0].price_role == "review_candidate"


def test_review_prefers_higher_score_then_lower_price():
    item = CanonicalItem(
        id="y",
        sheet="询价",
        row=4,
        name="LED地埋灯",
        spec="9W DC24V",
        submit=258.0,
    )
    attempts = [
        {
            "platform": "zaojiatong",
            "price_tax": 270.0,
            "match_ok": False,
            "match_score": 0.9,
            "match_detail": "名称命中；规格缺少：DC24V",
            "title": "LED地埋灯",
            "name_hit": True,
            "tax_mode": "tax_excl",
            "url": "https://example.com/near",
        },
        {
            "platform": "zaojiatong",
            "price_tax": 200.0,
            "match_ok": False,
            "match_score": 0.5,
            "match_detail": "名称命中；规格缺少：DC24V",
            "title": "LED地埋灯",
            "name_hit": True,
            "tax_mode": "tax_excl",
            "url": "https://example.com/under",
        },
    ]
    review = build_review_candidates(item, attempts, limit=2)
    assert len(review) == 2
    # 分更高的 270 排前（价格不决定匹配，只在同分时价低优先）
    assert review[0].price == 270.0
    assert review[0].match_score >= review[1].match_score


def test_review_export_keeps_candidate_separate_and_links_exact_evidence(
    tmp_path: Path,
):
    """待核价不得回填审定价；来源规格/价证据/正确附件必须单独可核。"""
    source = tmp_path / "source.xlsx"
    output = tmp_path / "result.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "询价"
    ws.append(["材料名称", "规格型号"])
    ws.append(["不锈钢卡箍", "DN150"])
    wb.save(source)

    item = CanonicalItem(
        id="询价|2",
        sheet="询价",
        row=2,
        name="不锈钢卡箍",
        spec="DN150",
        submit=10,
    )
    evidence_url = "https://evidence.example.com/supplier-quote.pdf"
    review = Quote(
        rank=1,
        price=17.08,
        platform="guangcai",
        title="不锈钢卡箍",
        url="https://www.gldjc.com/scj/so.html?keyword=DN150",
        detail_url=evidence_url,
        match_level="need_review",
        match_detail="名称命中；规格缺少：尺寸 DN150",
        spec_seen="材质:不锈钢 | 规格(mm):150 | 类型:铸铁管用",
        unit="个",
        supplier="某供应商",
        contact="张经理",
        phone="13800000000",
        price_text="17.08",
        price_context="广材搜索结果第1个材料组 / 第2条厂家报价",
        evidence_scope="exact_quote_row",
        source_group_index=1,
        source_quote_index=2,
        source_row_label="广材搜索结果第1个材料组 / 第2条厂家报价",
        price_role="review_candidate",
    )
    qset = QuoteSet(
        item_id=item.id,
        review_candidates=[review],
        status="need_review",
    )
    restored_review = Quote.from_dict(review.to_dict())
    assert restored_review.source_group_index == 1
    assert restored_review.source_quote_index == 2
    assert "第2条厂家报价" in restored_review.source_row_label
    write_quote_result_workbook(
        source,
        output,
        [item],
        {item.id: qset},
        k=1,
    )

    out = openpyxl.load_workbook(output)
    summary = out["询价比价结果"]
    headers = {
        str(summary.cell(4, col).value): col
        for col in range(1, summary.max_column + 1)
    }
    assert summary.cell(5, headers["参考审定不含税"]).value is None
    assert summary.cell(5, headers["待核候选价"]).value == 17.08
    assert "规格(mm):150" in summary.cell(5, headers["来源规格"]).value
    assert "第2条厂家报价" in summary.cell(5, headers["来源报价行位置"]).value
    link_cell = summary.cell(5, headers["外部证据链接"])
    assert link_cell.value == "打开供应商报价附件"
    assert link_cell.hyperlink.target == evidence_url

    detail = out["待审核候选明细"]
    detail_headers = {
        str(detail.cell(4, col).value): col
        for col in range(1, detail.max_column + 1)
    }
    assert detail.cell(5, detail_headers["询价材料名称"]).value == "不锈钢卡箍"
    assert detail.cell(5, detail_headers["来源品名"]).value == "不锈钢卡箍"
    assert "规格(mm):150" in detail.cell(5, detail_headers["来源规格"]).value
    assert detail.cell(5, detail_headers["来源价格原文"]).value == "17.08"
    exact_link = detail.cell(5, detail_headers["外部证据链接"])
    assert exact_link.value == "打开供应商报价附件"
    assert exact_link.hyperlink.target == evidence_url
