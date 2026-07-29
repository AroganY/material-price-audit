"""Excel input discovery shared by the web wizard and diagnostics."""

from __future__ import annotations

from pathlib import Path

import openpyxl


# openpyxl supports OOXML workbooks, not the legacy binary .xls format.
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}


def r2(value: float) -> float:
    """Round a monetary value to two decimals without binary-edge surprises."""
    return round(float(value) + 1e-12, 2)


def _inquiry_content_score(path: Path) -> int:
    """Score a workbook by whether its first rows resemble an inquiry sheet."""
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return 0

    score = 0
    try:
        for sheet_name in workbook.sheetnames[:8]:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(max_row=12, max_col=22, values_only=True):
                text = " ".join(str(cell) for cell in row if cell is not None)
                if not text:
                    continue
                if any(key in text for key in ("报送不含税单价", "报送单价", "投标单价")):
                    score += 100
                if any(key in text for key in ("审定不含税单价", "审定单价")):
                    score += 20
                if any(key in text for key in ("材料名称", "设备名称", "规格、型号", "规格型号")):
                    score += 8
                if "数量" in text and "单位" in text:
                    score += 4
    finally:
        workbook.close()
    return score


def resolve_inquiry_path(
    path: str | Path | None,
    default_dir: Path | None = None,
) -> Path:
    """Resolve a workbook path without requiring a fixed filename or headers."""
    if path in (None, ""):
        if default_dir is None:
            raise FileNotFoundError("未指定询价表，且没有默认 data/input 目录")
        candidate = Path(default_dir)
    else:
        candidate = Path(path).expanduser()

    if candidate.is_file():
        if candidate.suffix.lower() not in EXCEL_SUFFIXES:
            raise FileNotFoundError(f"不是支持的 Excel 文件: {candidate}")
        if candidate.name.startswith("~$"):
            raise FileNotFoundError(f"不能读取 Excel 临时锁文件: {candidate}")
        return candidate.resolve()

    folder = candidate if candidate.suffix.lower() not in EXCEL_SUFFIXES else candidate.parent
    if not folder.exists() and default_dir and Path(default_dir).exists():
        folder = Path(default_dir)
    if not folder.is_dir():
        raise FileNotFoundError(f"路径不存在: {candidate}")

    workbooks = [
        workbook
        for suffix in EXCEL_SUFFIXES
        for workbook in folder.glob(f"*{suffix}")
        if not workbook.name.startswith(("~$", "."))
    ]
    if not workbooks:
        raise FileNotFoundError(f"目录中没有 Excel 文件: {folder}")

    name_hints = ("询价", "材料", "设备", "安装", "核价", "认价", "inquiry", "price", "audit")
    output_hints = ("result", "rfq", "证据", "核价完成", "审定核价", "output")

    def rank(workbook: Path) -> tuple[int, float]:
        name = workbook.name.lower()
        hint_score = sum(3 for hint in name_hints if hint.lower() in name)
        output_penalty = sum(15 for hint in output_hints if hint.lower() in name)
        return _inquiry_content_score(workbook) + hint_score - output_penalty, workbook.stat().st_mtime

    chosen = max(workbooks, key=rank).resolve()
    if len(workbooks) > 1:
        print(f"[input] 发现 {len(workbooks)} 个 Excel，按表头与文件名选用: {chosen.name}")
    return chosen
