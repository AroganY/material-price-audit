"""
Table structure understanding:
  L0 rule header mapping (no fixed single header names)
  L1 optional LLM schema mapping
  Cache by fingerprint for repeat templates
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import threading
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from .models import ColumnMap, SheetSchema, WorkbookSchema
from .settings_store import UserSettings

# role -> many possible header fragments (not exclusive fixed headers)
ROLE_HINTS: dict[str, list[str]] = {
    "name": [
        "材料名称", "设备名称", "材料设备名称", "材料及设备名称", "名称", "品名",
        "项目名称", "材料设备", "物资名称", "商品名称", "清单名称", "工程材料",
    ],
    "spec": [
        "规格、型号", "规格型号", "规格及型号", "规格", "型号", "技术参数",
        "规格参数", "特征描述", "项目特征", "材质规格",
    ],
    "brand": [
        "产地、品牌及特殊要求", "产地品牌及特殊要求", "产地、品牌", "品牌产地",
        "品牌", "产地", "厂家", "生产厂家", "厂商", "特殊要求",
    ],
    "unit": ["计量单位", "单位", "单位名称"],
    "qty": ["工程量", "数量", "清单工程量", "申报数量", "合同数量"],
    "submit_price": [
        "报送不含税单价", "报送单价", "投标单价", "承包人报价", "承包人申报单价",
        "报审单价", "申报单价", "报出单价", "不含税单价", "含税单价", "单价",
        "综合单价", "材料单价", "设备单价", "市场单价", "信息价",
    ],
    "audit_price": [
        "审定不含税单价", "审定单价", "核定单价", "审核单价", "批准单价", "认定单价",
    ],
    "sum_price": ["合价", "报送合价", "审定合价", "金额", "总价"],
    "remark": ["备注", "说明", "附注", "注释"],
    "region": [
        "地区",
        "所在地区",
        "项目地区",
        "工程所在地",
        "工程地点",
        "建设地点",
        "省市",
        "城市",
        "区域",
        "适用地区",
        "项目城市",
        "地区名称",
    ],
}

# higher = more specific when multiple match
ROLE_PRIORITY = {
    "submit_price": 50,
    "audit_price": 45,
    "name": 40,
    "spec": 38,
    "brand": 30,
    "qty": 28,
    "unit": 26,
    "region": 24,
    "sum_price": 20,
    "remark": 10,
    "ignore": 0,
    "unknown": 0,
}


def _cell_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\n", "").replace("\r", "").strip()


def fingerprint_workbook(path: Path, max_rows: int = 12, max_cols: int = 22) -> str:
    """Structure fingerprint from header-ish rows (stable across data edits)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for sn in wb.sheetnames[:12]:
            if sn.strip() in ("实抓汇总", "询价比价结果", "核价汇总", "说明", "README"):
                continue
            ws = wb[sn]
            parts.append(f"SHEET:{sn}")
            for i, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True)):
                cells = [_cell_text(c) for c in row]
                if any(cells):
                    parts.append(f"R{i+1}:" + "|".join(cells))
    finally:
        wb.close()
    raw = "\n".join(parts)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def cache_dir(root: Path) -> Path:
    d = root / "data" / "mapping-cache"
    d.mkdir(parents=True, exist_ok=True)
    return d


def load_schema_cache(root: Path, fp: str) -> WorkbookSchema | None:
    p = cache_dir(root) / f"{fp}.json"
    if not p.exists():
        return None
    try:
        return WorkbookSchema.from_dict(json.loads(p.read_text(encoding="utf-8")))
    except Exception:
        return None


def save_schema_cache(root: Path, schema: WorkbookSchema) -> Path:
    p = cache_dir(root) / f"{schema.file_fingerprint}.json"
    schema.created_at = schema.created_at or datetime.now().isoformat(timespec="seconds")
    p.write_text(json.dumps(schema.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    return p


def _score_header_text(text: str) -> tuple[str | None, float]:
    """Return best role + confidence for a header cell."""
    t = _cell_text(text)
    if not t or len(t) > 40:
        return None, 0.0
    best_role = None
    best_score = 0.0
    for role, hints in ROLE_HINTS.items():
        for h in hints:
            if h == t:
                sc = 1.0
            elif h in t or t in h:
                sc = 0.85 if len(h) >= 2 else 0.5
            else:
                continue
            # penalize bare「单价」if longer more specific already
            if h == "单价" and t != "单价":
                sc = 0.55
            if h == "名称" and t != "名称" and "材料" not in t and "设备" not in t and "项目" not in t:
                sc = min(sc, 0.7)
            pri = ROLE_PRIORITY.get(role, 0) / 100.0
            sc = sc + pri * 0.05
            if sc > best_score:
                best_score = sc
                best_role = role
    return best_role, min(1.0, best_score)


def _row_header_score(cells: list[str]) -> float:
    score = 0.0
    roles = set()
    for t in cells:
        role, sc = _score_header_text(t)
        if role and sc >= 0.55:
            score += sc
            roles.add(role)
    # bonus if looks like a real header row
    if "name" in roles:
        score += 1.5
    if "spec" in roles:
        score += 0.8
    if "submit_price" in roles or "unit" in roles:
        score += 0.8
    return score


def map_sheet_by_rules(ws, sheet_name: str, max_scan: int = 15, max_cols: int = 24) -> SheetSchema | None:
    grid: list[list[str]] = []
    for r in range(1, max_scan + 1):
        row = []
        for c in range(1, max_cols + 1):
            row.append(_cell_text(ws.cell(r, c).value))
        grid.append(row)

    best_r = None
    best_score = 0.0
    for i, row in enumerate(grid):
        sc = _row_header_score(row)
        if sc > best_score:
            best_score = sc
            best_r = i + 1  # 1-based

    if best_r is None or best_score < 1.2:
        return None

    header = grid[best_r - 1]
    columns: list[ColumnMap] = []
    used_roles: set[str] = set()
    # rank candidates then assign unique roles
    cands: list[tuple[float, int, str, str]] = []
    for c, text in enumerate(header, 1):
        if not text:
            continue
        role, conf = _score_header_text(text)
        if role and conf >= 0.55:
            cands.append((conf + ROLE_PRIORITY.get(role, 0) / 1000.0, c, role, text))
        else:
            columns.append(ColumnMap(col=c, role="unknown", header_text=text, confidence=0.3))

    cands.sort(reverse=True)
    assigned_cols: set[int] = set()
    for conf, c, role, text in cands:
        if c in assigned_cols:
            continue
        # allow only one of each primary role (except sum_price can be multi — take first)
        if role in used_roles and role != "sum_price":
            # secondary: if name already taken and text looks like remark
            columns.append(ColumnMap(col=c, role="ignore", header_text=text, confidence=conf * 0.5))
            assigned_cols.add(c)
            continue
        used_roles.add(role)
        assigned_cols.add(c)
        columns.append(ColumnMap(col=c, role=role, header_text=text, confidence=conf))

    columns.sort(key=lambda x: x.col)
    roles = {c.role for c in columns}
    if "name" not in roles:
        # try: first long text column as name
        for c, text in enumerate(header, 1):
            if text and c not in assigned_cols:
                columns.append(ColumnMap(col=c, role="name", header_text=text, confidence=0.4))
                roles.add("name")
                break
        if "name" not in roles:
            # column 1 or 2 often name
            for guess in (2, 1, 3):
                if guess <= max_cols:
                    columns.append(ColumnMap(col=guess, role="name", header_text=header[guess - 1], confidence=0.35))
                    break

    conf = min(1.0, best_score / 6.0)
    return SheetSchema(
        sheet=sheet_name.strip(),
        header_row=best_r,
        data_start_row=best_r + 1,
        columns=columns,
        layout_notes="rule-mapped",
        source="rule",
        confidence=conf,
    )


def _sheet_preview(ws, max_rows: int = 18, max_cols: int = 18) -> list[list[str]]:
    out = []
    for r in range(1, max_rows + 1):
        row = [_cell_text(ws.cell(r, c).value)[:40] for c in range(1, max_cols + 1)]
        if any(row):
            out.append(row)
    return out


def resolve_llm_api_key(settings: UserSettings) -> str:
    """优先本机向导保存的 Key，其次环境变量。"""
    key = (getattr(settings, "llm_api_key", None) or "").strip()
    if key:
        return key
    env_name = (settings.llm_api_key_env or "").strip()
    if env_name:
        key = (os.environ.get(env_name) or "").strip()
        if key:
            return key
    key = (os.environ.get("MATERIAL_PRICE_AUDIT_LLM_KEY") or "").strip()
    if key:
        return key
    for env in ("OPENAI_API_KEY", "SPACEXAI_API_KEY", "LLM_API_KEY"):
        key = (os.environ.get(env) or "").strip()
        if key:
            return key
    return ""


# 可选：记录每次 LLM usage（prompt/completion/total tokens）
_LLM_USAGE_HOOK = None  # type: ignore
_LLM_CALL_GUARD = None  # type: ignore


class LLMItemCallBudget:
    """单条材料跨平台共享的 API 调用预算（线程安全）。"""

    def __init__(self, max_calls: int = 1) -> None:
        self.max_calls = max(0, int(max_calls or 0))
        self.calls = 0
        self._lock = threading.Lock()

    def reserve(self, req: dict[str, Any] | None = None) -> dict[str, Any]:
        with self._lock:
            if self.calls >= self.max_calls:
                return {
                    "allowed": False,
                    "reason": f"本条材料 AI 调用已达 {self.max_calls} 次上限",
                }
            self.calls += 1
            return {"allowed": True, "call_index": self.calls}


def set_llm_usage_hook(hook) -> None:
    """hook(usage: dict) — usage 含 prompt_tokens/completion_tokens/total_tokens/model/ok"""
    global _LLM_USAGE_HOOK
    _LLM_USAGE_HOOK = hook


def set_llm_call_guard(hook) -> None:
    """设置 API 请求前硬预算钩子；返回 False/allowed=False 时不发请求。"""
    global _LLM_CALL_GUARD
    _LLM_CALL_GUARD = hook


def _estimate_tokens(*parts: str) -> int:
    """API 未返回 usage 时粗估（约 4 字符 ≈ 1 token，中英混合够用）。"""
    n = sum(len(p or "") for p in parts)
    return max(1, (n + 3) // 4)


def _emit_llm_usage(
    *,
    ok: bool,
    model: str = "",
    prompt_tokens: int = 0,
    completion_tokens: int = 0,
    total_tokens: int = 0,
    role: str = "",
) -> None:
    if not _LLM_USAGE_HOOK:
        return
    try:
        _LLM_USAGE_HOOK(
            {
                "ok": bool(ok),
                "model": model or "",
                "role": role or "",
                "prompt_tokens": int(prompt_tokens or 0),
                "completion_tokens": int(completion_tokens or 0),
                "total_tokens": int(
                    total_tokens
                    or (int(prompt_tokens or 0) + int(completion_tokens or 0))
                ),
            }
        )
    except Exception:
        pass


def _parse_usage_tokens(usage: dict[str, Any], system: str, user: str, content: str = "") -> tuple[int, int, int]:
    """兼容 OpenAI / 部分代理的 usage 字段名。"""
    usage = usage or {}
    pt = int(
        usage.get("prompt_tokens")
        or usage.get("input_tokens")
        or usage.get("prompt_token_count")
        or 0
    )
    ct = int(
        usage.get("completion_tokens")
        or usage.get("output_tokens")
        or usage.get("completion_token_count")
        or 0
    )
    tt = int(usage.get("total_tokens") or usage.get("total_token_count") or 0)
    if not tt:
        pt = pt or _estimate_tokens(system, user)
        ct = ct or (_estimate_tokens(content) if content else 0)
        tt = pt + ct
    return pt, ct, tt


def _llm_chat_json(
    settings: UserSettings,
    system: str,
    user: str,
    *,
    role: str = "chat",
) -> dict | None:
    if not settings.llm_enabled:
        return None
    key = resolve_llm_api_key(settings)
    if not key:
        print("[schema] LLM 已开启但未配置 API Key（向导填写或环境变量）")
        return None
    base = (
        settings.llm_api_base
        or os.environ.get("OPENAI_API_BASE")
        or "https://api.openai.com/v1"
    ).rstrip("/")
    url = f"{base}/chat/completions"
    model = settings.llm_model or "gpt-4o-mini"
    use_role = (role or "chat").strip() or "chat"
    estimated_prompt_tokens = _estimate_tokens(system, user)
    if _LLM_CALL_GUARD:
        try:
            verdict = _LLM_CALL_GUARD(
                {
                    "role": use_role,
                    "model": model,
                    "estimated_prompt_tokens": estimated_prompt_tokens,
                }
            )
            allowed = (
                bool(verdict.get("allowed"))
                if isinstance(verdict, dict)
                else bool(verdict)
            )
            if not allowed:
                reason = (
                    str(verdict.get("reason") or "已达到 AI 硬预算")
                    if isinstance(verdict, dict)
                    else "已达到 AI 硬预算"
                )
                print(f"[schema] LLM 请求已阻止（未消耗 Token）: {reason}")
                return None
        except Exception as e:
            # 预算组件异常时采取 fail-closed，避免保护失效后继续烧 Token。
            print(f"[schema] LLM 预算检查异常，请求已阻止（未消耗 Token）: {e}")
            return None
    item_guard = getattr(settings, "_llm_item_call_budget", None)
    if item_guard is not None:
        try:
            verdict = item_guard.reserve(
                {
                    "role": use_role,
                    "model": model,
                    "estimated_prompt_tokens": estimated_prompt_tokens,
                }
            )
            if not bool(verdict.get("allowed")):
                print(
                    "[schema] LLM 请求已阻止（未消耗 Token）: "
                    + str(verdict.get("reason") or "本条材料 AI 预算已用尽")
                )
                return None
        except Exception as e:
            print(f"[schema] 本条材料 AI 预算检查异常，请求已阻止: {e}")
            return None
    # 限制单次输出；match_review 只需一个很短的 JSON，不能让模型输出上千 Token。
    max_output_tokens = {
        "match_review": 220,
        "search_agent": 400,
        "schema": 900,
    }.get(use_role, 700)
    body = {
        "model": model,
        "temperature": 0.1,
        "max_tokens": max_output_tokens,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
    }
    req = urllib.request.Request(
        url,
        data=json.dumps(body).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {key}",
        },
        method="POST",
    )
    try:
        # 询价链路里 LLM 只是辅助；超时过长会拖死整条瀑布匹配
        with urllib.request.urlopen(req, timeout=25) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
        content = raw["choices"][0]["message"]["content"]
        if isinstance(content, list):
            content = "".join(
                (x.get("text") if isinstance(x, dict) else str(x)) for x in content
            )
        # strip markdown fences if any
        content = content.strip()
        if content.startswith("```"):
            content = re.sub(r"^```(?:json)?\s*", "", content)
            content = re.sub(r"\s*```$", "", content)
        usage = raw.get("usage") if isinstance(raw, dict) else None
        if not isinstance(usage, dict):
            usage = {}
        pt, ct, tt = _parse_usage_tokens(usage, system, user, content)
        _emit_llm_usage(
            ok=True,
            model=str(raw.get("model") or model),
            prompt_tokens=pt,
            completion_tokens=ct,
            total_tokens=tt,
            role=use_role,
        )
        return json.loads(content)
    except (urllib.error.URLError, urllib.error.HTTPError, KeyError, json.JSONDecodeError, TimeoutError) as e:
        print(f"[schema] LLM 调用失败，回退规则: {type(e).__name__}: {e}")
        _emit_llm_usage(
            ok=False,
            model=model,
            prompt_tokens=_estimate_tokens(system, user),
            completion_tokens=0,
            total_tokens=_estimate_tokens(system, user),
            role=use_role,
        )
        return None
    except Exception as e:
        print(f"[schema] LLM 异常，回退规则: {e}")
        _emit_llm_usage(
            ok=False,
            model=model,
            prompt_tokens=_estimate_tokens(system, user),
            completion_tokens=0,
            total_tokens=_estimate_tokens(system, user),
            role=use_role,
        )
        return None


def check_llm_readiness(
    settings: UserSettings | None,
    *,
    probe: bool = False,
) -> dict[str, Any]:
    """
    运行前 AI 可用性检查。

    - want_ai: 用户是否开启
    - usable: 配置层面是否足以发起请求（Key/用途）
    - live_ok: probe=True 时是否真连通
    - blockers: 人类可读阻塞原因
    """
    s = settings or UserSettings()
    want = bool(s.llm_enabled)
    use_for = [str(x) for x in (s.llm_use_for or []) if str(x).strip()]
    model = (s.llm_model or "gpt-4o-mini").strip()
    base = (
        (s.llm_api_base or "").strip()
        or os.environ.get("OPENAI_API_BASE")
        or "https://api.openai.com/v1"
    ).rstrip("/")
    key = resolve_llm_api_key(s)
    key_ready = bool(key)
    blockers: list[str] = []

    if not want:
        return {
            "ok": True,
            "want_ai": False,
            "usable": False,
            "live_ok": None,
            "probed": False,
            "key_ready": key_ready,
            "model": model,
            "api_base": base,
            "use_for": use_for,
            "blockers": [],
            "summary": "AI 未开启 · 将纯规则运行（正常）",
            "hint": "需要 AI 时请在第①步开启并配置 Key，再点「测试 AI 连接」。",
        }

    if not key_ready:
        blockers.append(
            "未配置 API Key（向导填写或环境变量 "
            f"{s.llm_api_key_env or 'OPENAI_API_KEY'}）"
        )
    if not use_for:
        blockers.append("未勾选任何 AI 用途（表头/语义/搜索）")
    known = {"schema", "match_review", "search_agent"}
    if use_for and not any(u in known for u in use_for):
        blockers.append("用途配置无效")

    usable = want and not blockers
    live_ok: bool | None = None
    probe_error = ""
    if probe and usable:
        # 强制视为开启做连通探测
        probe_settings = UserSettings(
            llm_enabled=True,
            llm_api_base=s.llm_api_base,
            llm_api_key_env=s.llm_api_key_env,
            llm_api_key=s.llm_api_key,
            llm_model=s.llm_model,
            llm_use_for=list(use_for) or ["schema"],
        )
        ping = test_llm_connection(probe_settings)
        live_ok = bool(ping.get("ok"))
        if not live_ok:
            probe_error = str(ping.get("error") or "连通测试失败")
            blockers.append(probe_error)

    if not usable:
        summary = "AI 已开启但不可用 · " + "；".join(blockers)
    elif probe and live_ok is False:
        summary = "AI 配置看似完整，但连通失败 · " + (probe_error or "请检查网络/Base/模型")
    elif probe and live_ok is True:
        summary = f"AI 可用 · 已探测连通 model={model}"
    else:
        summary = f"AI 配置可用 · model={model}（尚未探测连通，建议点「测试 AI 连接」）"

    return {
        "ok": usable and (live_ok is not False),
        "want_ai": True,
        "usable": usable,
        "live_ok": live_ok,
        "probed": bool(probe),
        "key_ready": key_ready,
        "model": model,
        "api_base": base,
        "use_for": use_for,
        "blockers": blockers,
        "summary": summary,
        "hint": (
            "请回到第①步：填写 Key → 勾选用途 → 点「测试 AI 连接」通过后再询价。"
            if blockers
            else "开始询价时会再确认；失败可选择改用纯规则继续。"
        ),
    }


def test_llm_connection(settings: UserSettings) -> dict[str, Any]:
    """向导「测试连接」：发一条最小 JSON 请求。"""
    # 测试时允许未勾「开启」——只要有 Key 就测
    if not resolve_llm_api_key(settings):
        return {
            "ok": False,
            "error": "未配置 API Key：请在下方填写，或设置环境变量 "
            f"{settings.llm_api_key_env or 'OPENAI_API_KEY'}",
        }
    # 临时开启以通过 _llm_chat_json 开关
    s = settings
    if not s.llm_enabled:
        from dataclasses import replace

        try:
            s = replace(settings, llm_enabled=True)
        except Exception:
            s.llm_enabled = True
    data = _llm_chat_json(
        s,
        '只输出 JSON：{"ok":true,"pong":"material-price-audit"}',
        "ping",
        role="ping",
    )
    if not data:
        return {
            "ok": False,
            "error": "调用失败：请检查 API Base / Key / 模型名是否正确，以及网络是否可达",
        }
    return {
        "ok": True,
        "message": f"连接成功（model={s.llm_model or 'gpt-4o-mini'}）",
        "sample": data,
    }


def map_sheet_by_llm(ws, sheet_name: str, settings: UserSettings) -> SheetSchema | None:
    if "schema" not in (settings.llm_use_for or ["schema"]):
        return None
    preview = _sheet_preview(ws)
    if not preview:
        return None
    system = (
        "你是工程造价询价表结构分析器。根据 Excel 网格识别表头行与列语义。"
        "只输出 JSON，不要定价。role 只能是: "
        "name,spec,brand,unit,qty,submit_price,audit_price,sum_price,remark,region,ignore,unknown。"
        "JSON 字段: header_row(int), data_start_row(int), columns:[{col,role,header_text,confidence}], "
        "layout_notes(string), confidence(0-1)。col 从 1 开始。"
    )
    user = json.dumps(
        {"sheet": sheet_name, "grid_rows": preview},
        ensure_ascii=False,
    )
    data = _llm_chat_json(settings, system, user, role="schema")
    if not data:
        return None
    cols = []
    for x in data.get("columns") or []:
        try:
            role = str(x.get("role") or "unknown")
            if role not in ROLE_HINTS and role not in ("ignore", "unknown"):
                role = "unknown"
            cols.append(
                ColumnMap(
                    col=int(x["col"]),
                    role=role,
                    header_text=str(x.get("header_text") or ""),
                    confidence=float(x.get("confidence") or 0.7),
                )
            )
        except Exception:
            continue
    if not cols:
        return None
    hr = int(data.get("header_row") or 1)
    return SheetSchema(
        sheet=sheet_name.strip(),
        header_row=hr,
        data_start_row=int(data.get("data_start_row") or (hr + 1)),
        columns=cols,
        layout_notes=str(data.get("layout_notes") or "llm"),
        source="llm",
        confidence=float(data.get("confidence") or 0.75),
    )


def detect_workbook_schema(
    path: Path,
    root: Path,
    settings: UserSettings | None = None,
    *,
    use_cache: bool = True,
    force_refresh: bool = False,
) -> WorkbookSchema:
    settings = settings or UserSettings()
    fp = fingerprint_workbook(path)
    if use_cache and not force_refresh:
        cached = load_schema_cache(root, fp)
        if cached and cached.sheets:
            print(f"[schema] 命中映射缓存 fingerprint={fp}")
            return cached

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets: list[SheetSchema] = []
    skip = {"实抓汇总", "询价比价结果", "核价汇总", "说明", "README", "待询价"}
    try:
        for sn in wb.sheetnames:
            if sn.strip() in skip:
                continue
            ws = wb[sn]
            schema = None
            # L0 rules first
            schema = map_sheet_by_rules(ws, sn)
            # L1 LLM if weak / no name
            need_llm = (
                schema is None
                or schema.confidence < 0.55
                or schema.role_col("name") is None
            )
            if need_llm and settings.llm_enabled:
                llm_schema = map_sheet_by_llm(ws, sn, settings)
                if llm_schema and (
                    schema is None
                    or llm_schema.confidence >= schema.confidence
                    or (schema.role_col("name") is None and llm_schema.role_col("name"))
                ):
                    schema = llm_schema
            if schema and schema.role_col("name"):
                sheets.append(schema)
                print(
                    f"[schema] sheet={sn!r} source={schema.source} "
                    f"header_row={schema.header_row} conf={schema.confidence:.2f} "
                    f"roles={list(schema.roles().keys())}"
                )
            else:
                print(f"[schema] sheet={sn!r} 无法识别有效材料列，跳过")
    finally:
        wb.close()

    result = WorkbookSchema(file_fingerprint=fp, sheets=sheets, created_at=datetime.now().isoformat(timespec="seconds"))
    if sheets:
        save_schema_cache(root, result)
    return result


def dump_schema_preview(schema: WorkbookSchema, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(schema.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    return path
