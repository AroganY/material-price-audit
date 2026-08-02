"""Export multi-quote inquiry results (side sheet + optional append cols)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .excel_io import r2
from .models import CanonicalItem, QuoteSet


def audit_from_quotes(
    item: CanonicalItem,
    qset: QuoteSet,
    tax_divisor: float,
    never_exceed: bool,
) -> float | None:
    if not qset.quotes:
        return None
    # prefer lowest ex-tax among quotes；忽略 0 / 占位
    prices = []
    for q in qset.quotes:
        if q.price is None or float(q.price) <= 0.05:
            continue
        ex = q.price_ex_tax
        if ex is not None and float(ex) <= 0.05:
            ex = None
        if ex is None and q.price and q.tax_mode == "tax_incl":
            ex = r2(q.price / tax_divisor)
        if ex is None and q.price:
            ex = r2(float(q.price))
        if ex is not None and float(ex) > 0.05:
            prices.append(ex)
    if not prices:
        return None
    audit = min(prices)
    if never_exceed and item.submit is not None:
        audit = min(audit, float(item.submit))
    return r2(audit)


def write_quote_result_workbook(
    source_path: Path,
    output_path: Path,
    items: list[CanonicalItem],
    quote_map: dict[str, QuoteSet],
    *,
    tax_divisor: float = 1.13,
    never_exceed: bool = True,
    k: int = 3,
    write_back_mode: str = "side_sheet",
) -> dict[str, int]:
    """
    write_back_mode:
      side_sheet — only 询价比价结果
      append_cols — also append 价1..K on source sheets
      both
    """
    wb = openpyxl.load_workbook(source_path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    partial_fill = PatternFill("solid", fgColor="FFF2CC")
    bad_fill = PatternFill("solid", fgColor="FCE4EC")
    link_font = Font(color="0563C1", underline="single")

    for dead in ("询价比价结果", "待审核候选明细", "实抓汇总"):
        if dead in wb.sheetnames:
            del wb[dead]

    summary = wb.create_sheet("询价比价结果", 0)
    summary["A1"] = f"询价比价结果（目标 {k} 个同名同规市场价；无名称规格匹配不编造）"
    summary["A1"].font = Font(bold=True, size=13, color="C00000")
    summary["A2"] = (
        f"不含税粗算=含税÷{tax_divisor}；"
        f"市场报价与报送价解耦：名称规格匹配即收录；"
        f"never_exceed_submit 仅影响「参考审定」封顶=min(最低市场不含税,报送)"
    )

    headers = [
        "专业/Sheet",
        "行号",
        "材料名称",
        "规格型号",
        "品牌",
        "单位",
        "数量",
        "报送单价",
        "目标地区",
        "名称规格匹配状态",
        "市场报价条数",
        "参考审定不含税",
    ]
    # 每个市场报价块列数（含与报送关系 + 地区）
    _QUOTE_COLS = 18
    for i in range(1, k + 1):
        headers.extend(
            [
                f"价{i}市场价格",
                f"价{i}不含税",
                f"价{i}税口径",
                f"价{i}平台",
                f"价{i}标题",
                f"价{i}厂家",
                f"价{i}联系人",
                f"价{i}电话",
                f"价{i}计价单位",
                f"价{i}起订量",
                f"价{i}价格证据",
                f"价{i}名称规格匹配",
                f"价{i}与报送关系",
                f"价{i}相对报送偏差%",
                f"价{i}异常提示",
                f"价{i}详情链接",
                f"价{i}价格适用地区",
                f"价{i}地区匹配",
            ]
        )
    _REVIEW_COLS = 13
    headers.extend(
        [
            "待核候选价",
            "候选平台",
            "来源品名",
            "来源规格",
            "来源单位",
            "来源厂家",
            "来源联系人",
            "来源电话",
            "来源价格原文",
            "来源报价行位置",
            "未满足规格",
            "证据类型",
            "外部证据链接",
            "电商参考价",
            "电商平台",
            "电商标题",
            "电商说明",
            "电商链接",
            "全网参考价",
            "全网来源域名",
            "全网标题",
            "全网质量",
            "全网说明",
            "全网链接",
            "供应商线索厂家",
            "供应商电话",
            "供应商联系人",
            "线索说明",
            "线索链接",
        ]
    )
    headers.append("解析备注")

    for c, h in enumerate(headers, 1):
        cell = summary.cell(4, c, h)
        cell.fill = header_fill
        cell.font = header_font

    stats = {
        "full_k": 0,
        "partial": 0,
        "need_review": 0,
        "no_match": 0,
        "other": 0,
        "items": len(items),
    }
    srow = 5
    for it in items:
        qset = quote_map.get(it.id) or quote_map.get(it.key) or QuoteSet(item_id=it.id, status="no_match")
        st = qset.status
        if st == "full_k":
            stats["full_k"] += 1
            fill = ok_fill
            st_label = f"已凑满{k}价"
        elif st == "partial":
            stats["partial"] += 1
            fill = partial_fill
            st_label = f"部分命中({len(qset.quotes)}/{k})"
        elif st == "need_review":
            stats["need_review"] += 1
            fill = partial_fill
            n_rev = len(qset.review_candidates or [])
            st_label = f"候选待核({n_rev}条)·请人工采用"
        elif st in ("no_match", "skipped"):
            stats["no_match"] += 1
            fill = bad_fill if st == "no_match" else partial_fill
            st_label = "没查到" if st == "no_match" else st
        else:
            stats["other"] += 1
            fill = partial_fill
            st_label = st

        audit = audit_from_quotes(it, qset, tax_divisor, never_exceed)
        n_q = len(qset.quotes)
        # 待核候选不是正式报价，绝不能回填「参考审定」冒充本材料已核价。
        hint = qset.error or ""
        if not n_q and qset.review_candidates:
            hint = hint or "有候选价待人工确认（非严格合格价）"
        elif not n_q:
            hint = hint or "无合格价且无候选"
        reg_disp = ""
        try:
            rd = getattr(it, "region", None) or {}
            if isinstance(rd, dict) and rd:
                reg_disp = "".join(
                    str(rd.get(k) or "")
                    for k in ("province", "city", "district")
                ) or str(rd.get("city") or "")
            if not reg_disp:
                reg_disp = str(getattr(it, "region_raw", "") or "")
        except Exception:
            reg_disp = ""
        # 报价上的目标地区（若有）
        if not reg_disp and qset.quotes:
            reg_disp = str(getattr(qset.quotes[0], "requested_region", "") or "")
        row_vals: list[Any] = [
            it.sheet,
            it.row,
            it.name[:60],
            it.spec[:80],
            it.brand[:40],
            it.unit,
            it.qty or None,
            it.submit,
            reg_disp[:40] if reg_disp else "",
            st_label,
            n_q,
            audit,
        ]
        for i in range(k):
            if i < n_q:
                q = qset.quotes[i]
                # 禁止导出 0 / 占位价（造价通锁价或解析失败时）
                q_price = q.price if (q.price is not None and float(q.price) > 0.05) else None
                ex = q.price_ex_tax
                if ex is not None and float(ex) <= 0.05:
                    ex = None
                if ex is None and q_price is not None and q.tax_mode == "tax_incl":
                    ex = r2(q_price / tax_divisor)
                detail = q.detail_url or q.url
                vs = getattr(q, "vs_submit", None) or "unknown"
                anomaly = getattr(q, "price_anomaly", None) or ""
                # 相对报送偏差%（不含税）
                dev_pct = None
                if (
                    it.submit is not None
                    and ex is not None
                    and float(it.submit) > 0
                ):
                    try:
                        dev_pct = r2(
                            (float(ex) - float(it.submit)) / float(it.submit) * 100.0
                        )
                    except Exception:
                        dev_pct = None
                vs_cn = {
                    "below_submit": "≤报送",
                    "near_submit": "接近报送",
                    "above_submit": "高于报送",
                    "suspicious_low": "异常偏低",
                    "unknown": "—",
                    "under": "≤报送",
                    "near": "接近报送",
                    "over": "高于报送",
                    "low": "异常偏低",
                }.get(str(vs), str(vs) or "—")
                row_vals.extend(
                    [
                        q_price,
                        ex,
                        q.tax_mode if q_price is not None else "",
                        q.platform,
                        (q.title or "")[:80],
                        (q.supplier or "")[:40],
                        (q.contact or "")[:20],
                        (q.phone or "")[:20],
                        (q.unit or "")[:12],
                        (q.moq or "")[:20],
                        (
                            getattr(q, "source_row_label", "")
                            or q.price_context
                            or q.price_text
                            or ""
                        )[:120],
                        f"{q.match_level}:{q.match_score:.2f}",
                        vs_cn,
                        dev_pct,
                        anomaly[:80] if anomaly else "",
                        detail,
                        (getattr(q, "source_price_region", None) or "")[:40],
                        (getattr(q, "region_match", None) or "")[:20],
                    ]
                )
            else:
                row_vals.extend([None] * _QUOTE_COLS)
        review = qset.review_candidates[0] if qset.review_candidates else None
        if review:
            rev_price = (
                review.price
                if (review.price is not None and float(review.price) > 0.05)
                else None
            )
            row_vals.extend(
                [
                    rev_price,
                    review.platform,
                    (review.title or "")[:80],
                    (review.spec_seen or "")[:500],
                    (review.unit or "")[:20],
                    (review.supplier or "")[:40],
                    (review.contact or "")[:40],
                    (review.phone or "")[:30],
                    (review.price_text or "")[:80],
                    (
                        getattr(review, "source_row_label", "")
                        or review.price_context
                        or ""
                    )[:160],
                    (review.match_detail or "")[:160],
                    (review.evidence_scope or "")[:80],
                    review.detail_url or review.url,
                ]
            )
        else:
            row_vals.extend([None] * _REVIEW_COLS)
        mref = (qset.market_refs or [None])[0]
        if mref:
            mref_price = (
                mref.price
                if (mref.price is not None and float(mref.price) > 0.05)
                else None
            )
            row_vals.extend(
                [
                    mref_price,
                    mref.platform,
                    (mref.title or "")[:80],
                    (mref.match_detail or "市场参考·非合格价")[:160],
                    mref.detail_url or mref.url,
                ]
            )
        else:
            row_vals.extend([None] * 5)
        wref = (getattr(qset, "web_refs", None) or [None])[0]
        if wref:
            w_price = (
                wref.price
                if (wref.price is not None and float(wref.price) > 0.05)
                else None
            )
            domain = ""
            try:
                from urllib.parse import urlparse

                host = urlparse(wref.detail_url or wref.url or "").netloc.lower()
                domain = host[4:] if host.startswith("www.") else host
            except Exception:
                domain = wref.platform or ""
            row_vals.extend(
                [
                    w_price,
                    domain or wref.platform,
                    (wref.title or "")[:80],
                    getattr(wref, "source_quality", "") or "",
                    (wref.match_detail or "全网参考·不进合格价")[:160],
                    wref.detail_url or wref.url,
                ]
            )
        else:
            row_vals.extend([None] * 6)
        slead = (getattr(qset, "supplier_leads", None) or [None])[0]
        if slead:
            row_vals.extend(
                [
                    (slead.supplier or "")[:40],
                    (slead.phone or "")[:20],
                    (slead.contact or "")[:20],
                    (slead.match_detail or "供应商线索·无可靠公开价")[:160],
                    slead.detail_url or slead.url,
                ]
            )
        else:
            row_vals.extend([None] * 5)
        note = "; ".join(it.parse_issues) if it.parse_issues else ""
        if hint:
            note = (note + " | " if note else "") + hint
        if qset.market_refs:
            note = (note + " | " if note else "") + f"电商参考{len(qset.market_refs)}条(不作合格价)"
        n_web = len(getattr(qset, "web_refs", None) or [])
        n_sup = len(getattr(qset, "supplier_leads", None) or [])
        if n_web:
            note = (note + " | " if note else "") + f"全网参考{n_web}条(不作合格价)"
        if n_sup:
            note = (note + " | " if note else "") + f"供应商线索{n_sup}条"
        row_vals.append(note)

        for c, v in enumerate(row_vals, 1):
            cell = summary.cell(srow, c, v)
            cell.fill = fill
        # hyperlinks：价块从第 13 列起（前 12 列为材料信息，含目标地区）
        _mat_cols = 12
        for i in range(k):
            # 详情链接在块内倒数第 3 列（后两列为价格适用地区、地区匹配）
            url_col = _mat_cols + 1 + i * _QUOTE_COLS + (_QUOTE_COLS - 3)
            if i < n_q and (qset.quotes[i].detail_url or qset.quotes[i].url):
                cell = summary.cell(srow, url_col)
                link = qset.quotes[i].detail_url or qset.quotes[i].url
                cell.hyperlink = link
                cell.value = "打开来源证据"
                cell.font = link_font
        if review and (review.detail_url or review.url):
            review_url_col = (
                _mat_cols + 1 + k * _QUOTE_COLS + (_REVIEW_COLS - 1)
            )
            cell = summary.cell(srow, review_url_col)
            cell.hyperlink = review.detail_url or review.url
            cell.value = (
                "打开供应商报价附件"
                if review.evidence_scope == "exact_quote_row"
                else "打开候选来源"
            )
            cell.font = link_font
        srow += 1

    summary.cell(
        srow + 1,
        1,
        f"full_k={stats['full_k']} partial={stats['partial']} "
        f"need_review={stats['need_review']} no_match={stats['no_match']} items={stats['items']}",
    )

    # 每个待审核候选单独占一行，避免汇总表只展示第一个候选、无法逐条核验。
    review_sheet = wb.create_sheet("待审核候选明细", 1)
    review_sheet["A1"] = "待审核候选明细（仅供人工核验，不是正式报价，不写入参考审定）"
    review_sheet["A1"].font = Font(bold=True, size=13, color="C00000")
    review_sheet["A2"] = (
        "询价材料字段与平台来源字段分开显示；优先打开供应商报价附件/材料详情，"
        "搜索页仅作为检索路径保留。"
    )
    review_headers = [
        "专业/Sheet",
        "行号",
        "询价材料名称",
        "询价规格型号",
        "候选序号",
        "待核候选价",
        "候选平台",
        "来源品名",
        "来源规格",
        "来源单位",
        "来源厂家",
        "来源联系人",
        "来源电话",
        "来源价格原文",
        "来源报价行位置",
        "未满足规格/待核原因",
        "证据类型",
        "外部证据链接",
        "搜索页",
    ]
    for c, h in enumerate(review_headers, 1):
        cell = review_sheet.cell(4, c, h)
        cell.fill = header_fill
        cell.font = header_font
    review_row = 5
    for it in items:
        qset = quote_map.get(it.id) or quote_map.get(it.key)
        if not qset:
            continue
        for index, candidate in enumerate(qset.review_candidates or [], 1):
            candidate_price = (
                candidate.price
                if candidate.price is not None and float(candidate.price) > 0.05
                else None
            )
            exact_url = candidate.detail_url or candidate.url or ""
            search_url = (
                candidate.url
                if candidate.url and candidate.url != candidate.detail_url
                else ""
            )
            values = [
                it.sheet,
                it.row,
                it.name,
                it.spec,
                index,
                candidate_price,
                candidate.platform,
                candidate.title or "",
                candidate.spec_seen or "",
                candidate.unit or "",
                candidate.supplier or "",
                candidate.contact or "",
                candidate.phone or "",
                candidate.price_text or "",
                (
                    getattr(candidate, "source_row_label", "")
                    or candidate.price_context
                    or ""
                ),
                candidate.match_detail or "",
                candidate.evidence_scope or "",
                exact_url,
                search_url,
            ]
            for c, value in enumerate(values, 1):
                review_sheet.cell(review_row, c, value).fill = partial_fill
            if exact_url:
                cell = review_sheet.cell(review_row, 18)
                cell.hyperlink = exact_url
                cell.value = (
                    "打开供应商报价附件"
                    if candidate.evidence_scope == "exact_quote_row"
                    else "打开候选来源"
                )
                cell.font = link_font
            if search_url:
                cell = review_sheet.cell(review_row, 19)
                cell.hyperlink = search_url
                cell.value = "打开搜索页"
                cell.font = link_font
            review_row += 1

    review_sheet.freeze_panes = "A5"
    for col, width in {
        "A": 14,
        "B": 8,
        "C": 28,
        "D": 34,
        "E": 10,
        "F": 14,
        "G": 12,
        "H": 26,
        "I": 52,
        "J": 12,
        "K": 28,
        "L": 24,
        "M": 18,
        "N": 16,
        "O": 34,
        "P": 42,
        "Q": 18,
        "R": 22,
        "S": 14,
    }.items():
        review_sheet.column_dimensions[col].width = width

    if write_back_mode in ("append_cols", "both"):
        _append_quote_cols(wb, items, quote_map, k=k, tax_divisor=tax_divisor)

    # widths
    for c in range(1, min(len(headers), 20) + 1):
        summary.column_dimensions[get_column_letter(c)].width = 12
    summary.column_dimensions["C"].width = 28
    summary.column_dimensions["D"].width = 24
    summary.freeze_panes = "A5"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return stats


def _append_quote_cols(
    wb,
    items: list[CanonicalItem],
    quote_map: dict[str, QuoteSet],
    *,
    k: int,
    tax_divisor: float,
) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    link_font = Font(color="0563C1", underline="single")
    by_sheet: dict[str, list[CanonicalItem]] = {}
    for it in items:
        by_sheet.setdefault(it.sheet, []).append(it)

    for sn, sheet_items in by_sheet.items():
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        # find header row from first item cols or scan
        hr = None
        for it in sheet_items:
            # use row-1 if looks like header? better: min row - 1
            hr = max(1, min(x.row for x in sheet_items) - 1)
            break
        if not hr:
            continue
        real_max = 1
        for c in range(1, 40):
            if ws.cell(hr, c).value is not None:
                real_max = max(real_max, c)
        base = real_max + 1
        labels = ["询价状态", "参考审定"]
        for i in range(1, k + 1):
            labels.extend([f"询价价{i}", f"询价平台{i}", f"询价链接{i}"])
        for i, lab in enumerate(labels):
            cell = ws.cell(hr, base + i, lab)
            cell.fill = header_fill
            cell.font = header_font

        for it in sheet_items:
            qset = quote_map.get(it.id) or QuoteSet(item_id=it.id)
            audit = audit_from_quotes(it, qset, tax_divisor, True)
            ws.cell(it.row, base, qset.status)
            ws.cell(it.row, base + 1, audit)
            for i in range(k):
                col = base + 2 + i * 3
                if i < len(qset.quotes):
                    q = qset.quotes[i]
                    q_price = (
                        q.price
                        if (q.price is not None and float(q.price) > 0.05)
                        else None
                    )
                    ws.cell(it.row, col, q_price)
                    ws.cell(it.row, col + 1, q.platform)
                    cell = ws.cell(it.row, col + 2, "打开")
                    if q.url:
                        cell.hyperlink = q.url
                        cell.font = link_font


def export_rfq_from_quotes(
    items: list[CanonicalItem],
    quote_map: dict[str, QuoteSet],
    output_path: Path,
    *,
    k: int = 3,
) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "待询价"
    headers = [
        "专业",
        "行号",
        "材料名称",
        "规格型号",
        "品牌",
        "单位",
        "数量",
        "报送单价",
        "状态",
        "已有价条数",
        "备注",
        "线索厂家",
        "线索电话",
        "线索联系人",
        "线索链接",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    n = 0
    r = 2
    for it in items:
        qset = quote_map.get(it.id) or QuoteSet(item_id=it.id, status="no_match")
        if len(qset.quotes) >= k and qset.status == "full_k":
            continue
        n += 1
        ws.cell(r, 1, it.sheet)
        ws.cell(r, 2, it.row)
        ws.cell(r, 3, it.name)
        ws.cell(r, 4, it.spec)
        ws.cell(r, 5, it.brand)
        ws.cell(r, 6, it.unit)
        ws.cell(r, 7, it.qty)
        ws.cell(r, 8, it.submit)
        ws.cell(r, 9, qset.status)
        ws.cell(r, 10, len(qset.quotes))
        note = "; ".join(it.parse_issues)
        if qset.error:
            note = (note + " | " if note else "") + qset.error
        leads = list(getattr(qset, "supplier_leads", None) or [])
        if leads:
            note = (note + " | " if note else "") + f"供应商线索{len(leads)}条"
        ws.cell(r, 11, note)
        lead0 = leads[0] if leads else None
        if lead0:
            ws.cell(r, 12, (lead0.supplier or "")[:60])
            ws.cell(r, 13, (lead0.phone or "")[:30])
            ws.cell(r, 14, (lead0.contact or "")[:30])
            ws.cell(r, 15, lead0.detail_url or lead0.url or "")
        r += 1

    # 供应商线索明细表（可有正式价不足时的厂家/电话）
    ws2 = wb.create_sheet("供应商线索")
    h2 = [
        "专业",
        "行号",
        "材料名称",
        "规格型号",
        "厂家",
        "电话",
        "联系人",
        "来源标题",
        "说明",
        "链接",
        "来源质量",
    ]
    for c, h in enumerate(h2, 1):
        ws2.cell(1, c, h)
    r2i = 2
    for it in items:
        qset = quote_map.get(it.id) or QuoteSet(item_id=it.id, status="no_match")
        for lead in getattr(qset, "supplier_leads", None) or []:
            ws2.cell(r2i, 1, it.sheet)
            ws2.cell(r2i, 2, it.row)
            ws2.cell(r2i, 3, it.name)
            ws2.cell(r2i, 4, it.spec)
            ws2.cell(r2i, 5, (lead.supplier or "")[:60])
            ws2.cell(r2i, 6, (lead.phone or "")[:30])
            ws2.cell(r2i, 7, (lead.contact or "")[:30])
            ws2.cell(r2i, 8, (lead.title or "")[:100])
            ws2.cell(r2i, 9, (lead.match_detail or "")[:200])
            ws2.cell(r2i, 10, lead.detail_url or lead.url or "")
            ws2.cell(r2i, 11, getattr(lead, "source_quality", "") or "")
            r2i += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return n
