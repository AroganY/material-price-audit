"""Load rows via schema → CanonicalItem + search queries + spec tokens."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

from .matching import (
    collapse_cjk_spaces,
    extract_tokens,
    name_search_core,
    normalize_material_name,
    strip_geo_noise,
)
from .models import CanonicalItem, SheetSchema, WorkbookSchema


def _cell(ws, r: int, c: int | None) -> Any:
    if not c:
        return None
    return ws.cell(r, c).value


def _s(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\n", " ").replace("\r", " ").strip()


def _f(v: Any) -> float | None:
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("￥", "").replace("¥", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def split_name_spec(blob: str) -> tuple[str, str]:
    """If name/spec glued, try to split."""
    blob = _s(blob)
    if not blob:
        return "", ""
    # common separators
    for sep in ("|", "／", "/", "；", ";", "，"):
        if sep in blob:
            a, b = blob.split(sep, 1)
            if len(a.strip()) >= 2 and len(b.strip()) >= 1:
                return a.strip(), b.strip()
    # model-like tail
    m = re.search(
        r"(.+?)\s+((?:DN|φ|Φ)\s*\d{2,3}.*|(?:DS-|RG-|ST|HM-|JB-)[A-Za-z0-9/\-\.]+.*)$",
        blob,
        re.I,
    )
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return blob, ""


def peel_dims_into_spec(name: str, spec: str) -> tuple[str, str]:
    """
    名称栏常把截面/有效长度粘在一起：
      XZP100型片式消声器 1250X400 有效长度：1500
    → name=XZP100型片式消声器，spec 追加尺寸参数。
    """
    name, spec = _s(name), _s(spec)
    if not name:
        return name, spec
    moved: list[str] = []
    rest = name
    for pat in (
        r"\d+(?:\.\d+)?\s*[xX×*]\s*\d+(?:\.\d+)?(?:\s*[xX×*]\s*\d+(?:\.\d+)?){0,2}",
        r"(?:有效)?(?:长度|宽度|高度|厚度|深度)\s*[：:为]?\s*\d+(?:\.\d+)?\s*(?:mm|cm|m)?",
    ):
        for m in re.finditer(pat, rest, flags=re.I):
            moved.append(m.group(0).strip())
        rest = re.sub(pat, " ", rest, flags=re.I)
    rest = re.sub(r"\s+", " ", rest).strip(" ，,;；/-")
    if not moved:
        return name, spec
    extra = " ".join(moved)
    new_spec = f"{spec} {extra}".strip() if spec else extra
    return rest or name, new_spec


def _is_measurement_token(t: str) -> bool:
    return bool(
        re.fullmatch(
            r"(?i)(?:(?:AC|DC)?\d+(?:\.\d+)?(?:V|W|W/m|K)|IP\d{2})",
            re.sub(r"\s+", "", t or ""),
        )
    )


def _query_features(
    name: str, spec: str, brand: str, tokens: list[str] | None = None
) -> dict[str, Any]:
    """从名称/规格抽出各站共用的检索特征（型号、口径、关键参数）。"""
    name, spec, brand = _s(name), _s(spec), _s(brand)
    tokens = list(tokens or extract_tokens(f"{name} {spec} {brand}"))
    name_short = re.split(r"[（(【\[]", name)[0].strip() or name
    name_short = name_short[:28]
    name_core = name_search_core(name_short) or name_short
    # 保留名称里的「N端口/N路」身份（name_search_core 会剥掉，电商站反而需要）
    port_in_name = ""
    m_port = re.search(r"(\d+\s*(?:端口|路))", name_short)
    if m_port:
        port_in_name = re.sub(r"\s+", "", m_port.group(1))

    model = next(
        (
            t
            for t in tokens
            if not _is_measurement_token(t)
            and not re.fullmatch(r"(?i)(?:DN|PN|IP)\s*\d+(?:\.\d+)?", t or "")
            and not re.fullmatch(r"[φΦ]\s*\d+(?:\.\d+)?", t or "")
            and (
                re.match(
                    r"^(?:DS-|RG-|ST|iDS-|HM-|JB-|MS-|LRS-|GTYQ-|ZN-|WDZN-)", t, re.I
                )
                or re.match(r"^[A-Z]{1,5}\d{3,}[A-Z0-9\-]*$", t, re.I)
            )
        ),
        None,
    )
    if not model and spec:
        m = re.search(
            r"((?:DS-|RG-|iDS-|HM-|JB-|MS-|LRS-)[A-Z0-9/\-\.]+|[A-Z]{1,4}\d{3,}[A-Z0-9\-]*)",
            spec,
            re.I,
        )
        if (
            m
            and not _is_measurement_token(m.group(1))
            and not re.fullmatch(
                r"(?i)(?:DN|PN|IP)\s*\d+(?:\.\d+)?", m.group(1)
            )
        ):
            model = m.group(1)

    sizes = [
        t
        for t in tokens
        if t.upper().startswith("DN") or t.startswith("φ") or t.startswith("Φ")
    ]
    if not sizes and spec:
        m = re.search(r"(?:DN|φ|Φ)\s*\d{2,3}", spec, re.I)
        if m:
            sizes.append(re.sub(r"\s+", "", m.group(0)))
    # 截面尺寸 1250X400 / 630x400（消声器、风口等）；优先放进检索词
    blob_ns = f"{name} {spec}"
    for m in re.finditer(
        r"(?<!\d)(\d{2,5})\s*[xX×*]\s*(\d{2,5})(?:\s*[xX×*]\s*(\d{2,5}))?(?!\d)",
        blob_ns,
    ):
        parts = [m.group(1), m.group(2)] + ([m.group(3)] if m.group(3) else [])
        sz = "x".join(parts)
        if sz not in sizes and f"{parts[0]}×{parts[1]}" not in sizes:
            sizes.append(sz)
    # 有效长度单独作为检索辅助（不替代截面）
    m_len = re.search(r"有效长度\s*[：:为]?\s*(\d{3,5})", blob_ns)
    length_hint = m_len.group(1) if m_len else ""

    identity: list[str] = []  # 决定产品身份的词
    for word in (
        "脱机",
        "联机",
        "无线",
        "有线",
        "户外",
        "户内",
        "室外",
        "室内",
        "防水",
        "防雨",
        "防爆",
        "阻燃",
        "耐火",
        "明装",
        "暗装",
    ):
        if word in spec or word in name:
            identity.append(word)
    for pat in (r"\d+\s*端口", r"\d+\s*通道", r"\d+\s*路"):
        m = re.search(pat, f"{name} {spec}", re.I)
        if m:
            identity.append(re.sub(r"\s+", "", m.group(0)))

    electrical: list[str] = []
    for pat in (
        r"(?:AC|DC)\s*\d+(?:\.\d+)?\s*V",
        r"\d+(?:\.\d+)?\s*W\s*(?:[/／]\s*(?:m|米))?",
        r"IP\s*\d{2}",
        r"\d{3,5}\s*K",
    ):
        m = re.search(pat, spec, re.I)
        if m:
            electrical.append(re.sub(r"\s+", "", m.group(0)))

    # 名称里若带 LED/成品 等前缀，造价站常用「LED地埋灯」也能搜到
    name_led = ""
    if re.search(r"(?i)LED", name) and name_core and "LED" not in name_core.upper():
        name_led = f"LED{name_core}"

    brand_hint = brand
    if model:
        mu = model.upper()
        if (mu.startswith("DS-") or mu.startswith("IDS-")) and not brand_hint:
            brand_hint = "海康威视"
        if mu.startswith("RG-") and not brand_hint:
            brand_hint = "锐捷"

    return {
        "name": name,
        "spec": spec,
        "brand": brand,
        "brand_hint": brand_hint,
        "name_short": name_short,
        "name_core": name_core,
        "name_led": name_led,
        "model": model or "",
        "sizes": sizes,
        "length_hint": length_hint,
        "identity": identity,
        "electrical": electrical,
        "port_in_name": port_in_name,
        "tokens": tokens,
    }


def _query_part_key(part: str) -> str:
    """检索词片段归一化；用于识别 ``DN100 DN100`` 等重复锚点。"""
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", (part or "").lower())


def _is_query_anchor(part: str) -> bool:
    """型号/口径/电气参数可做包含去重，普通品名不做激进裁剪。"""
    raw = re.sub(r"\s+", "", part or "")
    return bool(
        re.fullmatch(r"(?i)(?:DN|PN|IP)\d+(?:\.\d+)?", raw)
        or re.fullmatch(r"[φΦ]\d+(?:\.\d+)?", raw)
        or re.fullmatch(r"(?i)(?:AC|DC)?\d+(?:\.\d+)?(?:V|W|K|MPA|KPA|A)", raw)
        or re.fullmatch(
            r"(?i)[A-Z]{1,10}[A-Z0-9]*[-_/\.][A-Z0-9][A-Z0-9\-_/\.]*",
            raw,
        )
        or re.fullmatch(r"(?i)[A-Z]{1,8}\d{2,}[A-Z0-9\-_/\.]*", raw)
    )


def normalize_search_query(query: str) -> str:
    """
    清理单个搜索词：
      - 中文内部空格折叠
      - 地名 / 信息价字样剥离
      - 重复型号/规格去重

    例：``DN100 DN100`` → ``DN100``；
    ``UQK-12液位计 UQK-12`` → ``UQK-12液位计``；
    ``成都 薄 壁 不锈钢管`` → ``薄壁不锈钢管``。
    """
    # 先折空格 + 去地名，再做锚点去重
    q0 = strip_geo_noise(collapse_cjk_spaces(query or ""))
    parts = [p for p in re.split(r"\s+", q0.strip()) if p]
    out: list[str] = []
    for part in parts:
        # 零件级再去一次地名（防止「成都市」残留）
        part = strip_geo_noise(part)
        if not part:
            continue
        key = _query_part_key(part)
        if not key:
            continue
        # 纯地名零件丢弃
        if re.fullmatch(r"(?:全国|本市|当地)", part):
            continue
        existing_keys = [_query_part_key(x) for x in out]
        if key in existing_keys:
            continue
        # 当前片段只是前面品名中已经包含的型号/口径，不再重复追加。
        if _is_query_anchor(part) and any(key in old for old in existing_keys):
            continue
        # 先出现纯型号，后出现「型号+品名」时保留信息更完整的后者。
        remove_indexes = [
            i
            for i, old in enumerate(out)
            if _is_query_anchor(old)
            and _query_part_key(old)
            and _query_part_key(old) in key
        ]
        for i in reversed(remove_indexes):
            out.pop(i)
        out.append(part)
    return " ".join(out)


def _dedupe_queries(queries: list[str], *, max_n: int = 6, max_len: int = 40) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for q in queries:
        q = normalize_search_query(q)
        if len(q) < 2 or len(q) > max_len:
            continue
        k = q.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(q)
        if len(out) >= max_n:
            break
    return out


# 造价信息站：按「材料品名库」检索，短词+口径/关键规格
_COST_PLATFORM_IDS = frozenset({"guangcai", "lingcai", "huixun", "yize", "zaojiatong"})
# 电商：按「商品标题」检索，品牌+型号+参数
_ECOM_PLATFORM_IDS = frozenset({"jd", "1688", "taobao", "tmall", "zkh", "suning"})


def build_cost_site_queries(name: str, spec: str, brand: str, tokens: list[str] | None = None) -> list[str]:
    """
    广材 / 领材 / 慧讯 / 易择 / 造价通 搜法（信息价/材料库）。

    **名称优先两阶段**（检索召回，不放宽正式匹配门禁）：
      1) **仅核心品名**（先扩召回同物结果集，禁止夹地名）
      2) 品名 + 口径/型号（二次精准，供规格匹配用）
      3) 少量身份/电气变体（预算内）
    禁止一上来就用「品名+一堆参数」把搜索结果弄脏。
    """
    # 名称粘尺寸时先剥开；再折空格/去地名
    name2, spec2 = peel_dims_into_spec(name, spec)
    name2 = normalize_material_name(name2) or collapse_cjk_spaces(name2)
    spec2 = collapse_cjk_spaces(spec2)
    f = _query_features(name2, spec2, brand, tokens)
    core = strip_geo_noise(f["name_core"] or "") or f["name_core"]
    short = strip_geo_noise(f["name_short"] or "") or f["name_short"]
    model = f.get("model") or ""
    sizes = list(f.get("sizes") or [])
    brand_h = (f.get("brand") or f.get("brand_hint") or "").strip()
    queries: list[str] = []

    # —— 1) 仅品名（最高优先，先拿同物结果集）——
    if core:
        queries.append(core)

    # —— 2) 硬规格二次精准（排在纯品名后，优先型号+截面 / 品名+口径）——
    if model and sizes:
        queries.append(f"{model} {sizes[0]}")
    if core and sizes:
        queries.append(f"{core} {sizes[0]}")
    if core and model:
        queries.append(f"{core} {model}")
    # 无 DN/型号的设备靠“身份词 + 硬电气参数”精准召回。
    # 例：分控器 脱机 8端口；线型灯 DC24V 18W/m IP65。
    if core and f.get("identity"):
        identity_bits = list(f["identity"][:3])
        queries.append(f"{core} {' '.join(identity_bits)}")
    if core and f.get("electrical"):
        electrical_bits = list(f["electrical"][:3])
        queries.append(f"{core} {' '.join(electrical_bits)}")
    if f.get("name_led"):
        led = strip_geo_noise(str(f["name_led"]))
        if led and led.lower() != (core or "").lower():
            queries.append(led)

    # —— 3) 短整名 / 身份 / 品牌（预算内）——
    if short and short.lower() not in {q.lower() for q in queries}:
        if len(short) <= 16 and short != core:
            queries.append(short)
    if core and f.get("identity"):
        queries.append(f"{core} {f['identity'][0]}")
    if core and f.get("port_in_name") and f["port_in_name"] not in (core or ""):
        queries.append(f"{core} {f['port_in_name']}")
    if brand_h and core:
        queries.append(f"{brand_h} {core}")

    # 召回变体：镀锌钢管 → 镀锌管（禁止动「不锈钢」→「不锈管」）
    if core and len(core) >= 4 and "钢" in core and "不锈钢" not in core:
        stripped = core.replace("钢", "", 1)
        if len(stripped) >= 2:
            queries.append(stripped)

    # 最多 4 个：少搜、快、Token 少；规格精度靠匹配门禁
    return _dedupe_queries(queries, max_n=4, max_len=40)


def platform_query_budget(
    platform_id: str,
    *,
    cost_max: int = 3,
    ecom_max: int = 2,
) -> int:
    """
    每站检索词预算。
    造价站默认 3（纯品名 + 品名规格）；电商保持较低（默认 2，上限 3）。
    """
    pid = (platform_id or "").strip().lower()
    if pid in _ECOM_PLATFORM_IDS:
        return max(1, min(3, int(ecom_max or 2)))
    # 造价站 / 未知站：名称优先，少搜快准
    n = int(cost_max or 3)
    return max(2, min(4, n))


def rule_requery_from_failures(
    name: str,
    spec: str,
    brand: str,
    tried_queries: list[str],
    fail_reasons: list[str],
    tokens: list[str] | None = None,
    *,
    max_n: int = 3,
) -> list[str]:
    """
    规则侧原因感知改词（AI 关闭/失败时的兜底）。
    失败原因关键字：名称未命中 / 型号 / DN / 尺寸 / 规格缺少。
    """
    name2, spec2 = peel_dims_into_spec(name, spec)
    f = _query_features(name2, spec2, brand, tokens)
    core = f.get("name_core") or ""
    short = f.get("name_short") or ""
    model = f.get("model") or ""
    sizes = list(f.get("sizes") or [])
    blob = "；".join(str(x) for x in (fail_reasons or []))
    tried = {re.sub(r"\s+", " ", (q or "")).strip().lower() for q in (tried_queries or [])}
    candidates: list[str] = []

    def _add(q: str) -> None:
        q = re.sub(r"\s+", " ", (q or "")).strip()
        if len(q) < 2 or len(q) > 40:
            return
        if q.lower() in tried:
            return
        if q not in candidates:
            candidates.append(q)

    dn_fail = bool(re.search(r"(?i)DN|口径|通径|φ|直径|尺寸", blob))
    model_fail = "型号" in blob
    name_fail = "名称未命中" in blob or "名称" in blob and "未命中" in blob
    missing = "缺少" in blob or "缺失" in blob or "未展示" in blob

    # DN/尺寸错误 → 强制带正确口径/截面
    if dn_fail or missing:
        if core and sizes:
            _add(f"{core} {sizes[0]}")
            _add(f"{sizes[0]} {core}")
        if model and sizes:
            _add(f"{model} {sizes[0]}")
        # 禁止纯口径搜索（如只搜 DN100）：召回噪声极大，也会制造大量重复询价。

    # 型号错误 → 品名+型号 / 纯型号
    if model_fail or missing:
        if core and model:
            _add(f"{core} {model}")
        if model:
            _add(model)

    # 名称未命中 → 短名 / 核心品名 / LED 变体
    if name_fail:
        if short:
            _add(short)
        if core:
            _add(core)
        if f.get("name_led"):
            _add(str(f["name_led"]))

    # 通用兜底：仍无新词时补 品名+关键规格
    if not candidates:
        if core and sizes:
            _add(f"{core} {sizes[0]}")
        if core and model:
            _add(f"{core} {model}")
        if core and f.get("identity"):
            _add(f"{core} {f['identity'][0]}")
        if short:
            _add(short)

    return candidates[: max(1, max_n)]


def build_ecommerce_queries(name: str, spec: str, brand: str, tokens: list[str] | None = None) -> list[str]:
    """
    京东 / 1688 搜法（商品标题）：
      1) 品牌 + 型号
      2) 型号
      3) 品牌 + 品名 + 关键参数
      4) 完整短名（8端口分控器 / LED地埋灯）
      5) 品名 + 身份词 + 电气参数
    电商标题靠品牌型号和卖点参数，不靠纯行业库短词。
    """
    f = _query_features(name, spec, brand, tokens)
    core = f["name_core"]
    short = f["name_short"]
    brand_h = f["brand_hint"] or f["brand"]
    queries: list[str] = []

    if brand_h and f["model"]:
        queries.append(f"{brand_h} {f['model']}")
    if f["model"]:
        queries.append(f["model"])
    if brand_h and core:
        tail = " ".join((f["identity"][:1] + f["electrical"][:2])[:3])
        queries.append(f"{brand_h} {core} {tail}".strip())
    if short:
        queries.append(short)
    if f["name_led"] and f["name_led"] != short:
        queries.append(f["name_led"])
    # 紧凑参数串：人在京东常搜「地埋灯 9W 24V IP67」
    if core:
        compact = " ".join(
            x for x in ([core] + f["identity"][:1] + f["electrical"][:3]) if x
        )
        queries.append(compact)
    if core and f["sizes"]:
        queries.append(f"{core} {f['sizes'][0]}")
    # 过宽的光杆品名放最后，仅作兜底
    if core and core.lower() not in {q.lower() for q in queries}:
        queries.append(core)

    return _dedupe_queries(queries, max_n=5, max_len=40)


def build_platform_queries(
    platform_id: str,
    name: str,
    spec: str,
    brand: str,
    tokens: list[str] | None = None,
) -> list[str]:
    """按平台检索习惯生成搜索词；禁止全站同一套 query。"""
    pid = (platform_id or "").strip().lower()
    if pid in _ECOM_PLATFORM_IDS:
        return build_ecommerce_queries(name, spec, brand, tokens)
    if pid in _COST_PLATFORM_IDS:
        return build_cost_site_queries(name, spec, brand, tokens)
    # 未知站：偏保守，用造价站策略（短词）
    return build_cost_site_queries(name, spec, brand, tokens)


def build_queries(name: str, spec: str, brand: str, tokens: list[str]) -> list[str]:
    """
    默认/预览用搜索词（跨站并集，供 Excel 预览与兼容旧逻辑）。
    真正抓取时请用 ``build_platform_queries(platform_id, ...)``。
    """
    # 合并两套策略，保持原始短名靠前
    f = _query_features(name, spec, brand, tokens)
    merged = []
    if f["name_short"]:
        merged.append(f["name_short"])
    merged.extend(build_cost_site_queries(name, spec, brand, tokens))
    merged.extend(build_ecommerce_queries(name, spec, brand, tokens))
    return _dedupe_queries(merged, max_n=6, max_len=40)


# 规格字段标签，不能当 must 命中词（否则「电源:DC24V」会强制要标题含「电源」）
_SPEC_LABEL_STOP = frozenset(
    {
        "电源",
        "功率",
        "色温",
        "角度",
        "防护",
        "防护等级",
        "控制",
        "控制方式",
        "规格",
        "型号",
        "品牌",
        "单位",
        "备注",
        "电压",
        "电流",
        "材质",
        "接口",
        "尺寸",
        "长度",
        "宽度",
        "高度",
        "厚度",
        "产地",
        "要求",
    }
)


def build_must(name: str, spec: str, tokens: list[str]) -> list[str]:
    must: list[str] = []
    for t in tokens:
        if re.match(r"^(?:DS-|RG-|ST|iDS-|HM-|JB-|MS-|LRS-)", t, re.I):
            must.append(t)
            parts = re.split(r"[-/]", t)
            for p in parts:
                if len(p) >= 4:
                    must.append(p)
        if t.upper().startswith("DN") or t.startswith("φ") or t.startswith("Φ"):
            must.append(t)
    # Chinese name chunks — 只用名称，不用规格里的字段标签
    for m in re.finditer(r"[\u4e00-\u9fff]{2,8}", name or ""):
        w = m.group(0)
        if w in ("材料", "设备", "名称", "规格", "型号", "不含税", "单价", "成品"):
            continue
        if w in _SPEC_LABEL_STOP:
            continue
        must.append(w)
        break
    # 规格侧只取身份词，不取「电源/功率」标签
    for word in ("脱机", "联机", "防水", "防爆", "阻燃", "耐火", "明装", "暗装"):
        if word in (spec or ""):
            must.append(word)
            break
    seen = set()
    out = []
    for x in must:
        k = x.lower()
        if k not in seen and len(x) >= 2 and x not in _SPEC_LABEL_STOP:
            seen.add(k)
            out.append(x)
    return out[:12]


def row_to_item(ws, schema: SheetSchema, r: int) -> CanonicalItem | None:
    roles = schema.roles()
    name_c = roles.get("name")
    if not name_c:
        return None
    name = _s(_cell(ws, r, name_c))
    spec = _s(_cell(ws, r, roles.get("spec")))
    brand = _s(_cell(ws, r, roles.get("brand")))
    unit = _cell(ws, r, roles.get("unit"))
    qty = _f(_cell(ws, r, roles.get("qty"))) or 0.0
    submit = _f(_cell(ws, r, roles.get("submit_price")))
    remark = _s(_cell(ws, r, roles.get("remark")))
    region_raw = _s(_cell(ws, r, roles.get("region")))

    # skip empty / section headers
    if not name and submit is None:
        return None
    if not name and spec:
        name, spec2 = split_name_spec(spec)
        if spec2:
            spec = spec2
    if name and not spec:
        n2, s2 = split_name_spec(name)
        if s2:
            name, spec = n2, s2
    # 名称粘尺寸/有效长度 → 挪到规格（消声器、风管附件极常见）
    name, spec = peel_dims_into_spec(name, spec)
    # 中文内部空格折叠 + 去地名：避免「薄 壁 管」「成都××」搞崩品名/搜索
    name = normalize_material_name(name) or collapse_cjk_spaces(name)
    spec = collapse_cjk_spaces(spec)
    if not name:
        return None
    # skip pure section titles (no digits, very short, no unit/qty/price)
    if (
        submit is None
        and qty == 0
        and not spec
        and len(name) <= 6
        and not re.search(r"\d", name)
    ):
        # still allow if looks like material
        if name.endswith(("工程", "分部", "合计", "小计", "说明", "专业")):
            return None

    issues: list[str] = []
    conf = 1.0
    status = "ok"
    if not spec:
        issues.append("无规格")
        conf -= 0.25
        status = "weak"
    if submit is None:
        issues.append("无报送单价")
        conf -= 0.05

    tokens = extract_tokens(f"{name} {spec} {brand}")
    queries = build_queries(name, spec, brand, tokens)
    must = build_must(name, spec, tokens)
    if not queries:
        issues.append("无法生成搜索词")
        status = "fail"
        conf = min(conf, 0.2)

    item_id = f"{schema.sheet}|{r}"
    region_dict: dict = {}
    if region_raw:
        try:
            from .region_gate import parse_region_text

            region_dict = parse_region_text(region_raw, source="excel_row").to_dict()
        except Exception:
            region_dict = {"city": region_raw, "source": "excel_row"}
    return CanonicalItem(
        id=item_id,
        sheet=schema.sheet,
        row=r,
        name=name,
        spec=spec,
        brand=brand,
        unit=unit if unit is not None else "",
        qty=qty,
        submit=submit,
        remark=remark,
        cols=roles,
        spec_tokens=tokens,
        search_queries=queries,
        must_match=must,
        parse_confidence=max(0.0, conf),
        parse_status=status,
        parse_issues=issues,
        region=region_dict,
        region_raw=region_raw,
    )


def load_canonical_items(
    path: Path,
    schema: WorkbookSchema,
    *,
    max_row_scan: int = 5000,
) -> list[CanonicalItem]:
    wb = openpyxl.load_workbook(path, data_only=True)
    items: list[CanonicalItem] = []
    try:
        by_name = {s.sheet: s for s in schema.sheets}
        for sn in wb.sheetnames:
            sch = by_name.get(sn.strip()) or by_name.get(sn)
            if not sch:
                continue
            ws = wb[sn]
            # 写回原表必须用真实 sheet 名（可能带尾部空格）
            real_sheet = sn
            start = sch.data_start_row
            end = min(ws.max_row or start, start + max_row_scan - 1)
            for r in range(start, end + 1):
                it = row_to_item(ws, sch, r)
                if it and it.parse_status != "fail":
                    it.sheet = real_sheet
                    it.id = f"{real_sheet}|{r}"
                    items.append(it)
    finally:
        wb.close()
    return items


def save_canonical_json(items: list[CanonicalItem], path: Path) -> None:
    import json

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps([i.to_dict() for i in items], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
